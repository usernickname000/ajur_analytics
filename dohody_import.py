# ============================================================
# dohody_import.py — Импорт мастер-файла бухгалтерии "Доходы ГГГГ.xlsx"
#
# Читает лист "ДОХОДЫ ГГГГ" (тот же макет, что и Доходы_2025.xlsx,
# на основе которого построен ACCOUNTING_ROWS в analytics.py) и:
#   1. сопоставляет строки файла с каноническими названиями статей;
#   2. для строк, которых нет в CRM (крм_projects == []), готовит
#      обновление external_income.json — БЕЗ ручной перепечатки;
#   3. для CRM-строк ничего не пишет, а только показывает цифру
#      бухгалтерии рядом — для сверки с выгрузкой CRM;
#   4. проверяет, что сумма найденных строк сходится с итоговыми
#      строками файла (Total Ads/Events/Прочие/Commercial Sales) —
#      если не сходится, парсинг что-то упустил;
#   5. ищет выбросы (месяц статьи в разы больше остальных месяцев
#      той же статьи) — типичный признак опечатки/ошибки в исходнике.
#
# Ничего не пишет само по себе: parse_workbook() только читает,
# apply_to_external_income() — единственная функция с побочным
# эффектом, и вызывается отдельно, осознанно.
# ============================================================

import argparse
import json
import os
import re
import shutil
from datetime import datetime

import openpyxl

from analytics import ACCOUNTING_ROWS

MONTHS = [f"{m:02d}" for m in range(1, 13)]

# Канонические статьи, отсутствующие в CRM — единственные, которые
# импортёр имеет право писать в external_income.json.
EXTERNAL_ROW_KEYS = {name for name, crm_projects in ACCOUNTING_ROWS if not crm_projects}
CRM_ROW_KEYS = {name for name, crm_projects in ACCOUNTING_ROWS if crm_projects}

SUBTOTAL_LABELS = {
    'Total Ads Sales', 'Total Events Sales',
    'Total Прочие доходы', 'Total Commercial Sales',
}

# Строки, которые НЕ являются самостоятельными статьями бюджета — это
# расшифровка/подмножество другой строки того же блока (проверено сверкой
# с Total Ads Sales: их добавление задваивает выручку). У них никогда нет
# своей колонки Total — это единственный надёжный маркер таких строк, но
# сам по себе недостаточен (см. MERGE_GROUPS ниже — там колонки Total тоже
# нет, но строки СЛЕДУЕТ учитывать).
DETAIL_SKIP_LABELS = {'ТГ', 'МАХ', 'ВК', 'Дзен/Пакеты'}

# Строки без своей колонки Total, которые в файле не сведены в отдельную
# итоговую строку, но по факту являются самостоятельными статьями —
# сверка с Total Ads Sales показывает, что их сумма (131 000 ₽) в точности
# закрывает разницу между найденными статьями и итогом файла. Объединяем
# их в одну строку с каноническим именем, под которым в ACCOUNTING_ROWS уже
# существует статья 'СММ Доктор' (CRM-строка, используется для сверки).
MERGE_GROUPS = {
    'СММ Доктор': {'Доктор ТГ', 'Доктор MAX', 'Доктор прочие СММ'},
}

# ── Сопоставление "как статья называется в файле бухгалтерии" → ──
# ── "каноническое название" (совпадает с ключами ACCOUNTING_ROWS  ──
# ── и external_income.json).                                      ──
# Список составлен по факту файла "Доходы 2026.xlsx". Если бухгалтерия
# в следующем году переименует строку — она попадёт в "unmatched"
# при импорте, и map нужно будет дополнить вручную (см. README).
RAW_LABEL_MAP = {
    # ── Реклама ФОНТАНКА (CRM, для сверки) ──
    'Fontanka.ru - баннерная реклама':                              'Fontanka.ru - баннерная реклама',
    'Fontanka.ru - мобильная реклама':                              'Fontanka.ru - мобильная реклама',
    'Fontanka.ru- ТЕКСТЫ (объединенная ячейка)':                    'Fontanka.ru - ТЕКСТЫ',
    'Fontanka.ru - ТЕКСТЫ':                                         'Fontanka.ru - ТЕКСТЫ',
    'IC Доходы ФОНТАНКА':                                           'IC Доходы ФОНТАНКА',
    'Fontanka.ru - НАТИВ.-спец.проекты (отдел СП)':                 'Fontanka.ru - НАТИВ-спецпроекты',
    'IC Доходы ФОНТАНКА спецпроекты':                               'IC Доходы ФОНТАНКА спецпроекты',
    'СММ доп+из нативных проектов':                                 'СММ Фонтанка',
    'IC Доходы ФОНТАНКА СММ':                                       'IC Доходы ФОНТАНКА СММ',
    # ── Реклама ФОНТАНКА (внешние — пишутся в external_income.json) ──
    'Программатик ФОНТАНКА ТГ':                                     'Программатик ФОНТАНКА ТГ',
    'Медийный бартер ФОНТАНКА':                                     'Медийный бартер ФОНТАНКА',
    'Программатик ФОНТАНКА':                                        'Программатик ФОНТАНКА',
    'E-com ФОНТАНКА':                                                'E-com ФОНТАНКА',
    'Рекомендтельные системы ФОНТАНКА':                             'Рекомендательные системы ФОНТАНКА',
    'Рекомендательные системы ФОНТАНКА':                            'Рекомендательные системы ФОНТАНКА',
    # ── Реклама ДОКТОР (CRM, для сверки) ──
    'doctorpiter.ru - реклама в десктоп (баннеры + тексты)':        'doctorpiter.ru - реклама в десктоп',
    'doctorpiter.ru - информ.услуги':                               'doctorpiter.ru - информ.услуги',
    'doctorpiter.ru - НАТИВ.-спец.проекты (отдел СП)':              'doctorpiter.ru - НАТИВ-спецпроекты',
    'IC Доходы ДОКТОР':                                             'IC Доходы ДОКТОР',
    # ── Реклама ДОКТОР (внешние) ──
    'Программатик ДОКТОР ТГ':                                       'Программатик ДОКТОР ТГ',
    'Программатик ДОКТОР':                                          'Программатик ДОКТОР',
    'E-com ДОКТОР':                                                  'E-com ДОКТОР',
    'Рекомендтельные системы ДОКТОР':                               'Рекомендательные системы ДОКТОР',
    'Рекомендательные системы ДОКТОР':                              'Рекомендательные системы ДОКТОР',
    # ── Мероприятия (CRM, для сверки) ──
    'Мероприятия Массовые. ЭВЕНТЫ':                                 'Мероприятия Массовые (ЭВЕНТЫ)',
    'IC Доходы ЭВЕНТЫ':                                             'IC Доходы ЭВЕНТЫ',
    'Мероприятия. КС. ФОНТАНКА':                                    'Мероприятия КС ФОНТАНКА',
    # ── Мероприятия (внешние) ──
    'IC Доходы КС. ФОНТАНКА':                                       'IC Доходы КС ФОНТАНКА',
    'IC Доходы КС ФОНТАНКА':                                        'IC Доходы КС ФОНТАНКА',
    'Медийный бартер ЭВЕНТЫ':                                       'Медийный бартер ЭВЕНТЫ',
    'Мероприятия. КС. Медицина. ДОКТОР':                            'Мероприятия КС Медицина ДОКТОР',
    # ── Прочие доходы (все внешние) ──
    'Выручка 47 (в план)':                                          'Выручка 47 (в план)',
    'Выручка 47 (закупка)':                                         'Выручка 47 (закупка)',
    'ФФ/АМ взаимозачет (схлопывается в "0" у экономиста)':          'ФФ/АМ взаимозачет',
    'ФФ/АМ взаимозачет':                                            'ФФ/АМ взаимозачет',
    'ИРИ/АНО':                                                      'ИРИ/АНО/Петроцентр',
    'ИРИ/АНО/Петроцентр':                                           'ИРИ/АНО/Петроцентр',
    # ── Корректировки (второй блок файла, но те же колонки) ──
    'Взаимозачет/Затраты':                                          'Взаимозачет/Затраты',
    'Корректировка скидки (комиссия ХШМ)':                          'Корректировка скидки (комиссия ХШМ)',
}


def _norm(s):
    s = '' if s is None else str(s)
    s = s.replace('\xa0', ' ').strip()
    s = re.sub(r'\s+', ' ', s)
    return s


class ParsedRow:
    __slots__ = ('raw_label', 'canonical', 'months', 'total', 'sheet_row')

    def __init__(self, raw_label, canonical, months, total, sheet_row):
        self.raw_label = raw_label
        self.canonical = canonical
        self.months = months          # {'01': руб или None, ...}
        self.total = total            # руб или None (значение из колонки Total файла)
        self.sheet_row = sheet_row

    def computed_total(self):
        return sum(v for v in self.months.values() if isinstance(v, (int, float)))


class ParsedWorkbook:
    def __init__(self, path):
        self.path = path
        self.year = None
        self.sheet_name = None
        self.rows = []            # list[ParsedRow] — только строки первичной таблицы (3..45)
        self.plan_group = {}      # {'01': руб, ...} из строки "План группы"
        self.subtotals = {}       # {'Total Ads Sales': ParsedRow, ...}
        self.unmatched = []       # list[ParsedRow] с canonical=None
        self.warnings = []        # список текстовых предупреждений


def _find_sheet(wb):
    for name in wb.sheetnames:
        if re.match(r'^ДОХОДЫ\s*\d{4}\s*$', name.strip()):
            return name
    # fallback: первый лист
    return wb.sheetnames[0]


def _find_header_row(ws, max_scan=10):
    """Ищет строку 'SALES' с датами месяцев в соседних ячейках."""
    for r in range(1, max_scan + 1):
        if _norm(ws.cell(r, 1).value).upper() == 'SALES':
            return r
    raise ValueError("Не найдена строка заголовка 'SALES' — формат файла не распознан.")


def parse_workbook(path, sheet_name=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    sn = sheet_name or _find_sheet(wb)
    ws = wb[sn]

    pw = ParsedWorkbook(path)
    pw.sheet_name = sn

    header_row = _find_header_row(ws)

    # Месяцы: первые 12 ячеек после подписи 'SALES', должны быть датами
    month_cols = list(range(2, 14))  # B..M
    dates = [ws.cell(header_row, c).value for c in month_cols]
    year_candidates = [d.year for d in dates if isinstance(d, datetime)]
    if not year_candidates:
        raise ValueError("Не удалось определить год — в шапке нет дат месяцев.")
    pw.year = max(set(year_candidates), key=year_candidates.count)

    total_col = None
    for c in range(14, ws.max_column + 1):
        if _norm(ws.cell(header_row, c).value).lower() == 'total':
            total_col = c
            break
    if total_col is None:
        total_col = 14  # колонка N по умолчанию

    r = header_row + 1
    seen_commercial_sales = False
    empty_streak = 0
    merge_accum = {name: {'months': {m: None for m in MONTHS}, 'rows': 0, 'sheet_row': None}
                   for name in MERGE_GROUPS}
    merge_source_of = {src: group for group, sources in MERGE_GROUPS.items() for src in sources}
    # RAW_LABEL_MAP намеренно мапит несколько написаний на один canonical
    # (опечатки, переименования). Если обе строки встретятся в одном файле —
    # складываем их месяцы вместо того, чтобы одна тихо перезаписала другую
    # (rows_by_canon/build_external_update ниже строят словарь по canonical).
    rows_by_canonical = {}

    while r <= ws.max_row and not seen_commercial_sales:
        label = _norm(ws.cell(r, 1).value)
        if not label:
            empty_streak += 1
            if empty_streak > 2:
                pw.warnings.append(
                    f"Таблица оборвалась на строке {r} (2 пустые строки подряд) "
                    f"раньше, чем встретилась 'Total Commercial Sales' — проверь файл."
                )
                break
            r += 1
            continue
        empty_streak = 0

        months = {}
        for i, m in enumerate(MONTHS):
            v = ws.cell(r, 2 + i).value
            months[m] = v if isinstance(v, (int, float)) else None
        total_val = ws.cell(r, total_col).value
        total_val = total_val if isinstance(total_val, (int, float)) else None
        months_sum = sum(v for v in months.values() if isinstance(v, (int, float)))
        has_any_month = any(v is not None for v in months.values())

        if label in SUBTOTAL_LABELS:
            prow = ParsedRow(label, None, months, total_val, r)
            pw.subtotals[label] = prow
            if label == 'Total Commercial Sales':
                seen_commercial_sales = True
            r += 1
            continue

        if label in DETAIL_SKIP_LABELS:
            pw.warnings.append(
                f"Строка {r} «{label}» — детализация/подмножество другой строки "
                f"этого блока (без своей колонки Total), в импорт не включается."
            )
            r += 1
            continue

        if label in merge_source_of:
            group = merge_source_of[label]
            acc = merge_accum[group]
            for m, v in months.items():
                if v is not None:
                    acc['months'][m] = (acc['months'][m] or 0) + v
            acc['rows'] += 1
            acc['sheet_row'] = acc['sheet_row'] or r
            r += 1
            continue

        # Эффективный итог: берём готовую колонку Total, а если она не
        # заполнена — считаем сами по месяцам (в файле такие строки есть,
        # напр. 'Выручка 47 (в план)', и сверка с Total Прочие доходы
        # показывает, что их обязательно нужно учитывать).
        effective_total = total_val if total_val is not None else (months_sum if has_any_month else None)

        canonical = RAW_LABEL_MAP.get(label)
        prow = ParsedRow(label, canonical, months, effective_total, r)

        if effective_total is None:
            pw.warnings.append(f"Строка {r} «{label}» — полностью пустая, пропущена.")
            r += 1
            continue

        if canonical is None:
            pw.unmatched.append(prow)
        else:
            existing = rows_by_canonical.get(canonical)
            if existing is None:
                pw.rows.append(prow)
                rows_by_canonical[canonical] = prow
            else:
                for m, v in months.items():
                    if v is not None:
                        existing.months[m] = (existing.months[m] or 0) + v
                if effective_total is not None:
                    existing.total = (existing.total or 0) + effective_total
                pw.warnings.append(
                    f"Строка {r} «{label}» смэтчилась на уже встретившуюся статью "
                    f"«{canonical}» (первая — «{existing.raw_label}») — суммы сложены, "
                    f"а не перезаписаны."
                )
                existing.raw_label = f"{existing.raw_label} + {label}"

        r += 1

    for group, acc in merge_accum.items():
        if acc['rows'] == 0:
            continue
        months_sum = sum(v for v in acc['months'].values() if v)
        prow = ParsedRow(
            ' + '.join(sorted(MERGE_GROUPS[group])), group,
            acc['months'], months_sum, acc['sheet_row'],
        )
        if group in EXTERNAL_ROW_KEYS or group in CRM_ROW_KEYS:
            existing = rows_by_canonical.get(group)
            if existing is None:
                pw.rows.append(prow)
                rows_by_canonical[group] = prow
            else:
                for m, v in acc['months'].items():
                    if v is not None:
                        existing.months[m] = (existing.months[m] or 0) + v
                existing.total = (existing.total or 0) + months_sum
                pw.warnings.append(
                    f"Группа «{prow.raw_label}» смэтчилась на уже встретившуюся статью "
                    f"«{group}» (первая — «{existing.raw_label}») — суммы сложены, а не перезаписаны."
                )
                existing.raw_label = f"{existing.raw_label} + {prow.raw_label}"
        else:
            pw.unmatched.append(prow)

    if not seen_commercial_sales:
        pw.warnings.append(
            "Не нашли строку 'Total Commercial Sales' — не могу проверить, "
            "что все статьи учтены. Проверь файл вручную."
        )

    # ── Ищем "Взаимозачет/Затраты" и "Корректировка скидки" во втором ──
    # ── (сверочном) блоке файла — они есть не в первичной таблице.    ──
    already = {row.canonical for row in pw.rows}
    for extra_label, canonical in [
        ('Взаимозачет/Затраты', 'Взаимозачет/Затраты'),
        ('Корректировка скидки (комиссия ХШМ)', 'Корректировка скидки (комиссия ХШМ)'),
    ]:
        if canonical in already:
            continue
        for rr in range(r, min(ws.max_row, r + 60) + 1):
            if _norm(ws.cell(rr, 1).value) == extra_label:
                months = {}
                for i, m in enumerate(MONTHS):
                    v = ws.cell(rr, 2 + i).value
                    months[m] = v if isinstance(v, (int, float)) else None
                total_val = ws.cell(rr, total_col).value
                total_val = total_val if isinstance(total_val, (int, float)) else None
                if total_val is not None:
                    pw.rows.append(ParsedRow(extra_label, canonical, months, total_val, rr))
                break

    # ── "План группы" — строка с 12 месячными числами, а не мини-таблица ──
    for rr in range(header_row + 1, min(ws.max_row, r + 40) + 1):
        if _norm(ws.cell(rr, 1).value) == 'План группы':
            months = {}
            filled = 0
            for i, m in enumerate(MONTHS):
                v = ws.cell(rr, 2 + i).value
                months[m] = v if isinstance(v, (int, float)) else None
                if months[m] is not None:
                    filled += 1
            if filled >= 10:
                pw.plan_group = months
                break

    return pw


def cross_check_subtotals(pw):
    """
    Сверяет сумму найденных строк с итоговыми строками файла
    (Total Ads Sales / Total Events Sales / Total Прочие доходы).
    Возвращает список текстовых расхождений (пусто, если всё сошлось).
    """
    from analytics import ACCOUNTING_BLOCKS

    issues = []
    rows_by_canon = {row.canonical: row for row in pw.rows}

    for block_name, member_keys in ACCOUNTING_BLOCKS.items():
        subtotal_row = pw.subtotals.get(block_name)
        if subtotal_row is None or subtotal_row.total is None:
            continue
        computed = sum(
            rows_by_canon[k].total for k in member_keys
            if k in rows_by_canon and rows_by_canon[k].total is not None
        )
        diff = computed - subtotal_row.total
        if abs(diff) > max(1.0, abs(subtotal_row.total) * 0.001):
            issues.append(
                f"{block_name}: сумма найденных строк = {computed:,.0f} ₽, "
                f"в файле указано {subtotal_row.total:,.0f} ₽ "
                f"(расхождение {diff:,.0f} ₽). Часть строк не сопоставлена "
                f"или пропущена без Total — см. unmatched/warnings."
            )
    return issues


def detect_outliers(pw, ratio=8.0):
    """
    Ищет месяцы статьи, которые в `ratio` раз больше медианы остальных
    ненулевых месяцев той же статьи — частый признак ошибки на порядок
    (как найденное +225 000 000 ₽ вместо -225 000 ₽ в этом же файле).
    """
    import statistics

    warnings = []
    for row in list(pw.rows) + list(pw.subtotals.values()):
        vals = [abs(v) for v in row.months.values() if isinstance(v, (int, float)) and v]
        if len(vals) < 3:
            continue
        for m, v in row.months.items():
            if not isinstance(v, (int, float)) or v == 0:
                continue
            others = [x for x in vals if x != abs(v)]
            if len(others) < 2:
                continue
            med = statistics.median(others)
            if med > 0 and abs(v) > med * ratio:
                warnings.append(
                    f"«{row.raw_label}», {m}.{pw.year}: {v:,.0f} ₽ — в {abs(v) / med:.0f}× "
                    f"больше медианы остальных месяцев этой строки ({med:,.0f} ₽). "
                    f"Похоже на опечатку/ошибку масштаба в исходнике — проверь у бухгалтерии."
                )
    return warnings


def build_external_update(pw):
    """
    Возвращает {canonical_key: {'01': руб|None, ...}} только для строк,
    которых нет в CRM (EXTERNAL_ROW_KEYS). CRM-строки сюда не попадают —
    их нельзя писать в external_income.json, иначе задвоится с выгрузкой CRM.
    """
    update = {}
    for row in pw.rows:
        if row.canonical in EXTERNAL_ROW_KEYS:
            update[row.canonical] = dict(row.months)
    return update


def build_crm_reference(pw):
    """Возвращает {canonical_key: {'01': руб|None, ...}} для CRM-строк — только для сверки."""
    return {row.canonical: dict(row.months) for row in pw.rows if row.canonical in CRM_ROW_KEYS}


def apply_to_external_income(json_path, pw, backup=True):
    """
    Обновляет external_income.json найденными внешними статьями.
    Пишет только те месяцы, где в файле есть число (не трогает месяцы,
    для которых в исходнике пусто — например, ещё не наступившие).
    При смене отчётного года делает резервную копию старого файла.
    Возвращает текстовый отчёт о том, что изменилось.
    """
    update = build_external_update(pw)

    if os.path.exists(json_path):
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    else:
        data = {
            "_описание": "Внешние доходы для бухгалтерской таблицы — того, чего нет в CRM.",
            "_включать_в_аналитику": [
                "Программатик ФОНТАНКА", "Программатик ФОНТАНКА ТГ",
                "Программатик ДОКТОР", "Программатик ДОКТОР ТГ",
            ],
            "_не_включать_в_grand_total": [
                "Выручка 47 (закупка)", "Взаимозачет/Затраты",
                "ФФ/АМ взаимозачет", "Корректировка скидки (комиссия ХШМ)",
            ],
        }

    old_year = data.get('_год')
    year_switch = bool(old_year) and old_year != pw.year
    report_lines = []

    if year_switch and backup:
        backup_path = os.path.join(
            os.path.dirname(json_path),
            f"external_income_{old_year}_backup.json",
        )
        if not os.path.exists(backup_path):
            shutil.copy2(json_path, backup_path)
            report_lines.append(f"Резервная копия за {old_year} год сохранена: {os.path.basename(backup_path)}")

    data['_год'] = pw.year

    for canonical, months in update.items():
        # При смене отчётного года не переносим цифры прошлого года на
        # месяцы, которых ещё нет в источнике — иначе под меткой 2026-го
        # тихо останутся факты 2025-го. Начинаем с чистого листа; при
        # обновлении в рамках ТОГО ЖЕ года — сохраняем то, что уже есть
        # (могло быть поправлено вручную через редактор).
        if year_switch:
            existing = {m: 0 for m in MONTHS}
        else:
            existing = data.get(canonical)
            if not isinstance(existing, dict):
                existing = {m: 0 for m in MONTHS}
        changed_months = []
        for m, v in months.items():
            if v is None:
                continue
            old_v = existing.get(m, 0) or 0
            if round(float(old_v), 2) != round(float(v), 2):
                changed_months.append(f"{m}: {old_v:,.0f} → {v:,.0f}")
            existing[m] = round(float(v), 2)
        data[canonical] = existing
        if changed_months:
            report_lines.append(f"{canonical}: " + "; ".join(changed_months))

    if pw.plan_group:
        plan = data.get('_план_группы_по_месяцам_руб', {})
        for m, v in pw.plan_group.items():
            if v is not None:
                plan[m] = round(float(v), 2)
        data['_план_группы_по_месяцам_руб'] = plan
        report_lines.append("Обновлён план группы по месяцам.")

    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    if not report_lines:
        report_lines.append("Изменений нет — external_income.json уже соответствует файлу.")

    return "\n".join(report_lines)


def generate_report(pw):
    lines = []
    lines.append(f"Файл: {os.path.basename(pw.path)}  |  Лист: {pw.sheet_name}  |  Год: {pw.year}")
    lines.append("")

    external_matched = [row for row in pw.rows if row.canonical in EXTERNAL_ROW_KEYS]
    crm_matched = [row for row in pw.rows if row.canonical in CRM_ROW_KEYS]

    lines.append(f"Статьи для external_income.json (будут записаны): {len(external_matched)}")
    for row in sorted(external_matched, key=lambda r: r.canonical):
        lines.append(f"  • {row.canonical}: {row.total:,.0f} ₽ за год (из файла)")

    lines.append("")
    lines.append(f"CRM-статьи (только для сверки, НЕ записываются): {len(crm_matched)}")
    for row in sorted(crm_matched, key=lambda r: r.canonical):
        lines.append(f"  • {row.canonical}: {row.total:,.0f} ₽ (бухгалтерия)")

    if pw.unmatched:
        lines.append("")
        lines.append(f"⚠ Не сопоставлено ({len(pw.unmatched)}) — нужно вручную дополнить RAW_LABEL_MAP:")
        for row in pw.unmatched:
            lines.append(f"  • строка {row.sheet_row}: «{row.raw_label}» = {row.total:,.0f} ₽")

    issues = cross_check_subtotals(pw)
    if issues:
        lines.append("")
        lines.append("⚠ Сумма статей не сходится с итоговыми строками файла:")
        for i in issues:
            lines.append(f"  • {i}")
    else:
        lines.append("")
        lines.append("✓ Сумма найденных статей сходится с Total Ads/Events/Прочие Sales.")

    outliers = detect_outliers(pw)
    if outliers:
        lines.append("")
        lines.append(f"⚠ Похожие на ошибку выбросы ({len(outliers)}):")
        for o in outliers:
            lines.append(f"  • {o}")

    if pw.warnings:
        lines.append("")
        lines.append("Прочие предупреждения:")
        for w in pw.warnings:
            lines.append(f"  • {w}")

    if pw.plan_group:
        plan_total = sum(v for v in pw.plan_group.values() if v)
        lines.append("")
        lines.append(f"План группы по месяцам найден, годовой итог {plan_total:,.0f} ₽.")

    return "\n".join(lines)


def main():
    ap = argparse.ArgumentParser(description="Импорт Доходы ГГГГ.xlsx в external_income.json")
    ap.add_argument("xlsx_path", help="Путь к файлу 'Доходы ГГГГ.xlsx'")
    ap.add_argument("--external", default=None,
                     help="Путь к external_income.json (по умолчанию — рядом со скриптом)")
    ap.add_argument("--apply", action="store_true",
                     help="Записать изменения (по умолчанию — только показать отчёт)")
    args = ap.parse_args()

    here = os.path.dirname(os.path.abspath(__file__))
    external_path = args.external or os.path.join(here, "external_income.json")

    pw = parse_workbook(args.xlsx_path)
    print(generate_report(pw))

    if args.apply:
        print("\n── Применяю изменения ──")
        print(apply_to_external_income(external_path, pw))
    else:
        print(f"\n(Режим предпросмотра. Чтобы записать в {external_path}, добавь --apply)")


if __name__ == "__main__":
    main()
