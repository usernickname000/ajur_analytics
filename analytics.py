# ============================================================
# analytics.py — ядро аналитики заказов
# Запускается из app.py, не трогать без необходимости
# ============================================================

import pandas as pd
import numpy as np
import os
import json
from datetime import datetime
from openpyxl.utils import get_column_letter


# ==============================
# ПАРАМЕТРЫ (можно редактировать)
# ==============================

EXCLUDE_MANAGERS = [
    'Алмазова А.', 'Василенко В.', 'Большакова М.',
    'Савиных Е.', 'Пирожкова Е.',
    'Селиванова А.', 'Дмитриева Д.', 'Матвеева В.'
]

EXCLUDE_PROJECTS = [
    'Мероприятия (деловые)',
    'Мероприятия (городские)'
]

DISCOUNT_BUCKETS = [0, 5, 10, 15, 20, 25, 50]

COL_BARTER = 'Бартер'
PROGRAMMATIC_PROJECTS = []

# ── Верифицированные цифры из бухгалтерии ──────────────────
# Загружаются из verified_figures.json при запуске.
# Константы-fallback на случай если JSON отсутствует:
VERIFIED_FIGURES_JSON = 'verified_figures.json'
VERIFIED_TOTAL_WITH_BARTER_NO_PROG = 363_000_166   # fallback
VERIFIED_ADVERTISING_NO_EVENTS     = 243_005_820   # fallback
VERIFIED_TOTAL_WITH_PROG           = 482_404_000   # fallback
EXTERNAL_PROGRAMMATIC_TOTAL        = 119_403_897   # fallback
EXTERNAL_OTHER_INCOME              = 41_057_102    # fallback

# ── Путь к JSON с внешними доходами по месяцам (для бухгалтерской таблицы) ──
# Если файл есть — строится полная бухгалтерская таблица с разбивкой по месяцам.
# Если файла нет или JSON пустой — лист '16_Бухгалтерская_таблица' будет пропущен.
EXTERNAL_INCOME_JSON = 'external_income.json'

# ── Маппинг строк бухгалтерии на CRM-проекты ────────────────────
# Формат: (название строки в бухгалтерии, список проектов CRM которые туда входят)
# Если список пуст — строка берётся только из EXTERNAL_INCOME_JSON (внешние данные).
# Порядок строк совпадает со структурой Доходы_2025.xlsx.
ACCOUNTING_ROWS = [
    # ── Блок: Реклама ФОНТАНКА ──
    ('Fontanka.ru - баннерная реклама',       ['Фонтанка.ру: Баннерная реклама']),
    ('Fontanka.ru - мобильная реклама',       ['Фонтанка.ру: Мобильная версия']),
    ('Fontanka.ru - ТЕКСТЫ',                  ['Фонтанка.ру: Тексты']),
    ('IC Доходы ФОНТАНКА',                    ['IC Доходы Фонтанка']),
    ('Fontanka.ru - НАТИВ-спецпроекты',       ['Фонтанка.ру: Спецпроекты']),
    ('IC Доходы ФОНТАНКА спецпроекты',        ['IC Доходы Фонтанка спецпроекты']),
    ('СММ Фонтанка',                          ['Фонтанка.ру: СММ']),
    ('IC Доходы ФОНТАНКА СММ',                ['IC Доходы Фонтанка СММ']),
    ('Программатик ФОНТАНКА ТГ',              []),  # внешний
    ('Медийный бартер ФОНТАНКА',              []),  # внешний
    ('Программатик ФОНТАНКА',                 []),  # внешний
    ('E-com ФОНТАНКА',                        []),  # внешний
    ('Рекомендательные системы ФОНТАНКА',     []),  # внешний
    # ── Блок: Реклама ДОКТОР ──
    ('doctorpiter.ru - реклама в десктоп',    ['ДокторПитер.ру: Реклама (баннеры + тексты)']),
    ('СММ Доктор',                            ['doctorpiter.ru: СММ']),
    ('doctorpiter.ru - информ.услуги',        ['ДокторПитер.ру: Информ. услуги']),
    ('doctorpiter.ru - НАТИВ-спецпроекты',    ['ДокторПитер.ру: Спецпроекты']),
    ('IC Доходы ДОКТОР',                      ['IC Доходы Доктор', 'IC Доходы Доктор (спецпроекты)']),
    ('Программатик ДОКТОР ТГ',                []),  # внешний
    ('Программатик ДОКТОР',                   []),  # внешний
    ('E-com ДОКТОР',                          []),  # внешний
    ('Рекомендательные системы ДОКТОР',       []),  # внешний
    # ── Блок: Мероприятия ──
    ('Мероприятия Массовые (ЭВЕНТЫ)',         ['Мероприятия (городские)']),
    ('Медийный бартер ЭВЕНТЫ',                []),  # внешний
    ('IC Доходы ЭВЕНТЫ',                      ['IC Доходы Эвенты']),
    ('Мероприятия КС ФОНТАНКА',               ['Мероприятия (деловые)']),
    ('Мероприятия КС Медицина ДОКТОР',        []),  # в CRM нет отдельно
    # ── Блок: Прочие доходы ──
    ('Выручка 47 (в план)',                   []),  # внешний — в бух часть 47 идёт отдельно
    ('Выручка 47 (закупка)',                  []),  # внешний
    ('ФФ/АМ взаимозачет',                     []),  # внешний
    ('Гранты Фонтанка',                       []),  # внешний
    ('Гранты Доктор',                         []),  # внешний
    ('ИРИ/АНО/Петроцентр',                    []),  # внешний
    # ── Корректировки ──
    ('Взаимозачет/Затраты',                   []),  # внешний
    ('Корректировка скидки (комиссия ХШМ)',   []),  # внешний (отрицательные суммы)
]

# Группировка строк бухгалтерии в блоки для промежуточных итогов
ACCOUNTING_BLOCKS = {
    'Total Ads Sales': [
        'Fontanka.ru - баннерная реклама', 'Fontanka.ru - мобильная реклама',
        'Fontanka.ru - ТЕКСТЫ', 'IC Доходы ФОНТАНКА',
        'Fontanka.ru - НАТИВ-спецпроекты', 'IC Доходы ФОНТАНКА спецпроекты',
        'СММ Фонтанка', 'IC Доходы ФОНТАНКА СММ',
        'Программатик ФОНТАНКА ТГ', 'Медийный бартер ФОНТАНКА',
        'Программатик ФОНТАНКА', 'E-com ФОНТАНКА', 'Рекомендательные системы ФОНТАНКА',
        'doctorpiter.ru - реклама в десктоп', 'СММ Доктор',
        'doctorpiter.ru - информ.услуги', 'doctorpiter.ru - НАТИВ-спецпроекты',
        'IC Доходы ДОКТОР', 'Программатик ДОКТОР ТГ', 'Программатик ДОКТОР',
        'E-com ДОКТОР', 'Рекомендательные системы ДОКТОР',
    ],
    'Total Events Sales': [
        'Мероприятия Массовые (ЭВЕНТЫ)', 'Медийный бартер ЭВЕНТЫ',
        'IC Доходы ЭВЕНТЫ', 'Мероприятия КС ФОНТАНКА', 'Мероприятия КС Медицина ДОКТОР',
    ],
    'Total Прочие доходы': [
        'Выручка 47 (в план)', 'Выручка 47 (закупка)', 'ФФ/АМ взаимозачет',
        'Гранты Фонтанка', 'Гранты Доктор', 'ИРИ/АНО/Петроцентр',
    ],
}

# Маппинг направлений для KPI
KPI_DIRECTIONS = {
    'Фонтанка': [
        'Fontanka.ru - баннерная реклама', 'Fontanka.ru - мобильная реклама',
        'Fontanka.ru - ТЕКСТЫ', 'IC Доходы ФОНТАНКА',
        'Fontanka.ru - НАТИВ-спецпроекты', 'IC Доходы ФОНТАНКА спецпроекты',
    ],
    'Доктор': [
        'doctorpiter.ru - реклама в десктоп', 'СММ Доктор',
        'doctorpiter.ru - информ.услуги', 'doctorpiter.ru - НАТИВ-спецпроекты',
        'IC Доходы ДОКТОР',
    ],
    'СММ': ['СММ Фонтанка', 'IC Доходы ФОНТАНКА СММ'],
    'Мероприятия + КС': [
        'Мероприятия Массовые (ЭВЕНТЫ)', 'IC Доходы ЭВЕНТЫ',
        'Мероприятия КС ФОНТАНКА', 'Мероприятия КС Медицина ДОКТОР',
    ],
}


# ── Маппинг проектов CRM → бизнес-группа ────────────────────────
# Структура повторяет логику Доходы_2025.xlsx: Ads / Events / Прочее / Программатик.
# Если появится новый проект — добавь его сюда, иначе он попадёт в "НЕ КЛАССИФИЦИРОВАНО".
PROJECT_GROUP_MAP = {
    # ── Реклама Фонтанка ──
    'Фонтанка.ру: Тексты':              'Реклама Фонтанка',
    'Фонтанка.ру: СММ':                 'Реклама Фонтанка',
    'Фонтанка.ру: Спецпроекты':         'Реклама Фонтанка',
    'Фонтанка.ру: Баннерная реклама':   'Реклама Фонтанка',
    'Фонтанка.ру: Мобильная версия':    'Реклама Фонтанка',
    'Фонтанка.ру':                      'Реклама Фонтанка',
    'IC Доходы Фонтанка':               'Реклама Фонтанка',
    'IC Доходы Фонтанка спецпроекты':   'Реклама Фонтанка',
    'IC Доходы Фонтанка СММ':           'Реклама Фонтанка',
    # ── Реклама Доктор ──
    'ДокторПитер.ру: Реклама (баннеры + тексты)': 'Реклама Доктор',
    'ДокторПитер.ру: Спецпроекты':      'Реклама Доктор',
    'doctorpiter.ru: СММ':              'Реклама Доктор',
    'ДокторПитер.ру: Информ. услуги':   'Реклама Доктор',
    'IC Доходы Доктор':                 'Реклама Доктор',
    'IC Доходы Доктор (спецпроекты)':   'Реклама Доктор',
    # ── Программатик (CRM-часть, малая) ──
    'Программатик Фонтанка':            'Программатик',
    'Программатик ДокторПитер':         'Программатик',
    # ── Мероприятия ──
    'Мероприятия (деловые)':            'Мероприятия',
    'Мероприятия (городские)':          'Мероприятия',
    'IC Доходы Эвенты':                 'Мероприятия',
    # ── 47News / Прочее ──
    '47News':                           '47News / Прочее',
    '47News: Прочие услуги':            '47News / Прочее',
}


REVENUE_BUCKETS = [
    ('Менее 10 тыс. руб.', 0, 10_000),
    ('10-50 тыс. руб.', 10_000, 50_000),
    ('50-100 тыс. руб.', 50_000, 100_000),
    ('100-500 тыс. руб.', 100_000, 500_000),
    ('500 тыс. - 1 млн руб.', 500_000, 1_000_000),
    ('Более 1 млн руб.', 1_000_000, float('inf'))
]

# ==============================
# ИМЕНА КОЛОНОК
# ==============================

COL_MONTH       = 'Месяц'
COL_ORDER       = 'Заказ'
COL_DATE        = 'Дата заказа'
COL_MANAGER     = 'Менеджер'
COL_REVENUE     = 'Выручка без НДС'
COL_DISCOUNT_PCT = 'Скидка, %'
COL_PROJECT     = 'Проект'

COL_CLIENT      = 'Клиент'
COL_CLIENT_RA   = 'Клиент РА'
COL_REKLAMD     = 'Рекламодатель'

COL_INDUSTRY_CLIENT_RA = 'Отрасли КлиентаРА'
COL_INDUSTRY_CLIENT    = 'Отрасли Клиента'
COL_INDUSTRY_ADV       = 'Отрасли Рекламодателя'

COL_DESCRIPTION = 'Описание'
COL_POS_ORDER   = 'Позиция заказа'
COL_NOMEN       = 'Номенклатура'


# ==============================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ==============================

def _rev_stats(df, groupby_col, revenue_col, avg_mask=None):
    """Groupby с суммой/количеством/средним чеком и переводом в тыс. руб."""
    grp = df.groupby(groupby_col)
    stats = grp[revenue_col].agg(['sum', 'count']).round(2)
    stats.columns = ['Сумма выручки, руб.', 'Количество заказов']
    avg_df = df if avg_mask is None else df.loc[avg_mask]
    stats['Средний чек, руб.'] = avg_df.groupby(groupby_col)[revenue_col].mean().round(2)
    stats['Сумма выручки, тыс. руб.'] = (stats['Сумма выручки, руб.'] / 1000).round(2)
    return stats.reset_index().sort_values('Сумма выручки, тыс. руб.', ascending=False)


def _find_rev_col(ws, fallback=2):
    """Ищет в заголовке листа колонку с выручкой в тыс. руб."""
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val and 'тыс' in str(val).lower() and 'выруч' in str(val).lower():
            return c
    return fallback


def parse_money(x):
    if pd.isna(x):
        return None
    x = str(x).strip()
    x = x.replace("руб", "").replace("Руб", "")
    x = x.replace(" ", "").replace("'", "")
    if "," in x and x.count(",") == 1 and x.count(".") >= 1:
        x = x.replace(".", "").replace(",", ".")
    else:
        x = x.replace(",", ".")
    try:
        return float(x)
    except:
        return None


def pick_client(row):
    for c in [COL_CLIENT_RA, COL_CLIENT, COL_REKLAMD]:
        if c in row and pd.notna(row[c]) and str(row[c]).strip() != '':
            return str(row[c]).strip()
    return ''


def normalize_client(name):
    if pd.isna(name):
        return name
    name_clean = str(name).upper()
    if 'ГАЗПРОМ' in name_clean:
        return 'ГАЗПРОМ'
    if 'АЛЬКОР' in name_clean or 'ГЛАВСТРОЙ' in name_clean:
        return 'АЛЬКОР / ГЛАВСТРОЙ'
    SBER_PATTERNS = [
        'СБЕР', 'SBER', 'СБЕРБАНК', 'СБЕР ЛИЗИНГ', 'СБЕРМАРКЕТ',
        'СБЕР ЗДОРОВЬЕ', 'СБЕРСТРАХ', 'СБЕРЛОГИСТИК',
        'ГЛОБАЛ МЕДИА ЕВРАЗИЯ', 'GLOBAL MEDIA EURASIA',
        'С-МАРКЕТИНГ', 'С МАРКЕТИНГ'
    ]
    if any(p in name_clean for p in SBER_PATTERNS):
        return 'ГРУППА СБЕР'
    return name


def pick_industry(row):
    for c in [COL_INDUSTRY_CLIENT_RA, COL_INDUSTRY_CLIENT, COL_INDUSTRY_ADV]:
        if c in row and pd.notna(row[c]) and str(row[c]).strip() != '':
            return str(row[c]).strip()
    return ''


def classify_industry(row):
    raw = str(row.get('ОТРАСЛЬ_КЛИЕНТА', '')).strip().lower()
    text = (
        str(row.get(COL_POS_ORDER, '') or '') + ' ' +
        str(row.get(COL_DESCRIPTION, '') or '') + ' ' +
        str(row.get(COL_NOMEN, '') or '') + ' ' +
        str(row.get('КОНЕЧНЫЙ_КЛИЕНТ', '') or '')
    ).lower()

    if any(x in raw for x in ['строит', 'строй', 'недвиж']):
        return 'Строительство + Недвижимость'
    if raw in ['маркетинг', 'реклама', 'выставки']:
        return raw.capitalize()

    expo_keywords = ['выстав', 'expo', 'expо', 'стенд', 'форум', 'конгресс', 'ярмарк', 'фестиваль', 'event', 'ивент']
    if any(k in text for k in expo_keywords):
        return 'Выставки'

    ads_keywords = [
        'реклам', 'размещ', 'баннер', 'промо', 'digital', 'диджитал',
        'программатик', 'programmatic', 'performance', 'перфоманс',
        'таргет', 'target', 'cpm', 'cpc', 'cpp', 'охват', 'показ',
        'спецпроект', 'интеграц', 'media', 'медиа'
    ]
    if any(k in text for k in ads_keywords):
        return 'Реклама'

    marketing_keywords = [
        'маркет', 'бренд', 'branding', 'pr', 'пиар', 'smm', 'контент',
        'исслед', 'аналит', 'стратег', 'консалт', 'коммуникац'
    ]
    if any(k in text for k in marketing_keywords):
        return 'Маркетинг'

    if 'маркетинг, реклама, выставки' in raw:
        return 'Маркетинг'

    return row.get('ОТРАСЛЬ_КЛИЕНТА', '')

def categorize_revenue_amount(v):
    if pd.isna(v):
        return 'Нет данных'
    try:
        v = float(v)
    except:
        return 'прочие'
    for name, lo, hi in REVENUE_BUCKETS:
        if lo <= v < hi:
            return name
    return 'прочие'


def bucket_discount(x):
    if pd.isna(x):
        return 'без скидки'
    try:
        xv = float(x)
    except:
        return 'прочие'
    if xv == 0:
        return 'без скидки'
    if xv in DISCOUNT_BUCKETS:
        return f'{int(xv)}%'
    if 40 <= xv <= 75:
        return '40–75% (нестандарты)'
    return 'прочие'


def load_verified_figures(path):
    """
    Загружает verified_figures.json. Возвращает dict с ключами:
    total_with_prog, total_with_barter_no_prog, advertising_no_events,
    programmatic_external, other_external_income, _год.
    Если файла нет — возвращает None.
    """
    if not os.path.exists(path):
        return None
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return None


def validate_analytics_consistency(verified_data, external_json_path, log=print):
    """
    Проверяет базовые несоответствия между источниками для аналитики.
    Сейчас валидирует:
    - совпадение '_год' в verified_figures.json и external_income.json.
    """
    verified_year = None
    external_year = None

    if isinstance(verified_data, dict):
        vy = verified_data.get('_год')
        if isinstance(vy, (int, float)):
            verified_year = int(vy)

    if external_json_path and os.path.exists(external_json_path):
        try:
            with open(external_json_path, 'r', encoding='utf-8') as f:
                ext = json.load(f)
            ey = ext.get('_год')
            if isinstance(ey, (int, float)):
                external_year = int(ey)
        except Exception:
            pass

    if verified_year and external_year and verified_year != external_year:
        log(
            "⚠ Несоответствие годов: "
            f"verified_figures.json={verified_year}, "
            f"external_income.json={external_year}. "
            "Проверьте, что оба файла относятся к одному отчетному году."
        )


def get_external_monthly_totals(external_json_path):
    """
    Читает external_income.json и возвращает dict {месяц: сумма_руб} — 
    сумму тех внешних доходов, которые явно перечислены в '_включать_в_аналитику'
    (программатик, бартер, ИРИ, гранты — т.е. реальная выручка, а не вычеты).
    Если списка '_включать_в_аналитику' нет — суммируются ВСЕ не-служебные строки.
    Если файла нет или он пустой — возвращает None.
    """
    if not os.path.exists(external_json_path):
        return None
    try:
        with open(external_json_path, 'r', encoding='utf-8') as f:
            ext = json.load(f)
    except Exception:
        return None

    months = [f"{m:02d}" for m in range(1, 13)]
    monthly_total = {m: 0.0 for m in months}

    # Если задан список — берём только его, иначе все не-служебные
    include_list = ext.get('_включать_в_аналитику')
    if include_list:
        rows_to_sum = [r for r in include_list if r in ext and isinstance(ext[r], dict)]
    else:
        rows_to_sum = [
            r for r, v in ext.items()
            if not r.startswith('_') and isinstance(v, dict)
        ]

    for row_name in rows_to_sum:
        row_data = ext[row_name]
        for m in months:
            v = row_data.get(m, 0) or 0
            if isinstance(v, (int, float)):
                monthly_total[m] += v

    return monthly_total


def get_full_external_total(external_json_path):
    """
    Суммирует внешние доходы из external_income.json для сверки с бухгалтерией.
    Исключает статьи из '_не_включать_в_grand_total' (закупка, взаимозачёты, затраты —
    это не реальный доход, а технические/расходные проводки).
    Возвращает итог в рублях.
    """
    if not os.path.exists(external_json_path):
        return 0.0
    try:
        with open(external_json_path, 'r', encoding='utf-8') as f:
            ext = json.load(f)
    except Exception:
        return 0.0

    # Статьи которые не являются реальным доходом и не должны попадать в grand_total
    exclude = set(ext.get('_не_включать_в_grand_total', [
        'Выручка 47 (закупка)',
        'Взаимозачет/Затраты',
        'ФФ/АМ взаимозачет',
        'Корректировка скидки (комиссия ХШМ)',
    ]))

    months = [f"{m:02d}" for m in range(1, 13)]
    total = 0.0
    for key, val in ext.items():
        if key.startswith('_') or not isinstance(val, dict):
            continue
        if key in exclude:
            continue
        for m in months:
            v = val.get(m, 0) or 0
            if isinstance(v, (int, float)):
                total += v
    return total


def get_external_totals_by_category(external_json_path):
    """
    Возвращает dict с разбивкой внешних доходов по категориям для сверки:
    {
        'programmatic': сумма программатика,
        'barter':       сумма бартера,
        'iri_grants':   сумма ИРИ/грантов,
        'recsys':       сумма рекомендательных систем,
        'ecom':         E-com,
        '47_plan':      47News в план (реальная выручка),
        'other':        остальные доходные статьи,
        'deductions':   взаимозачёты/затраты/корректировки (со знаком, обычно отриц.),
        'total_income': всё кроме deductions,
        'total_net':    total_income + deductions (итог для сверки),
    }
    """
    if not external_json_path or not os.path.exists(external_json_path):
        return {}
    try:
        with open(external_json_path, 'r', encoding='utf-8') as f:
            ext = json.load(f)
    except Exception:
        return {}

    months = [f"{m:02d}" for m in range(1, 13)]

    PROG_KEYS  = {'Программатик ФОНТАНКА', 'Программатик ФОНТАНКА ТГ',
                  'Программатик ДОКТОР', 'Программатик ДОКТОР ТГ'}
    BARTER_KEYS = {'Медийный бартер ФОНТАНКА', 'Медийный бартер ЭВЕНТЫ'}
    IRI_KEYS   = {'ИРИ/АНО/Петроцентр', 'Гранты Фонтанка', 'Гранты Доктор'}
    RECSYS_KEYS = {'Рекомендательные системы ФОНТАНКА', 'Рекомендательные системы ДОКТОР'}
    ECOM_KEYS   = {'E-com ФОНТАНКА', 'E-com ДОКТОР'}
    PLAN47_KEYS = {'Выручка 47 (в план)'}
    DEDUCT_KEYS = {'Выручка 47 (закупка)', 'Взаимозачет/Затраты',
                   'ФФ/АМ взаимозачет', 'Корректировка скидки (комиссия ХШМ)'}

    cats = {k: 0.0 for k in ('programmatic','barter','iri_grants','recsys',
                               'ecom','47_plan','other','deductions')}

    for key, val in ext.items():
        if key.startswith('_') or not isinstance(val, dict):
            continue
        s = sum((val.get(m, 0) or 0) for m in months if isinstance(val.get(m, 0), (int, float)))
        if key in PROG_KEYS:
            cats['programmatic'] += s
        elif key in BARTER_KEYS:
            cats['barter'] += s
        elif key in IRI_KEYS:
            cats['iri_grants'] += s
        elif key in RECSYS_KEYS:
            cats['recsys'] += s
        elif key in ECOM_KEYS:
            cats['ecom'] += s
        elif key in PLAN47_KEYS:
            cats['47_plan'] += s
        elif key in DEDUCT_KEYS:
            cats['deductions'] += s
        else:
            cats['other'] += s

    cats['total_income'] = sum(cats[k] for k in ('programmatic','barter','iri_grants',
                                                   'recsys','ecom','47_plan','other'))
    cats['total_net'] = cats['total_income'] + cats['deductions']
    return cats


def parse_month(s):
    try:
        if pd.isna(s):
            return None
        s = str(s).strip()
        if '.' in s:
            mm, yy = s.split('.')
            yy = int(yy)
            if yy < 100:
                yy = 2000 + yy
            return datetime(year=yy, month=int(mm), day=1)
        return pd.to_datetime(s, errors='coerce')
    except:
        return None


def build_accounting_table(df_full, revenue_col, external_json_path, log=print):
    """
    Строит бухгалтерскую таблицу в стиле Доходы_2025.xlsx.
    Возвращает DataFrame или None, если external_income.json не найден/пустой.
    """
    if not os.path.exists(external_json_path):
        log(f"⚠ Бухгалтерская таблица пропущена: нет файла {external_json_path}")
        return None

    try:
        with open(external_json_path, 'r', encoding='utf-8') as f:
            ext = json.load(f)
    except Exception as e:
        log(f"⚠ Не удалось прочитать {external_json_path}: {e}")
        return None

    # Месяцы: 01..12 как строки
    months = [f"{m:02d}" for m in range(1, 13)]
    analysis_year = int(ext.get('_год', 2025))
    month_labels = {f"{m:02d}": f"{m:02d}.{analysis_year}" for m in range(1, 13)}

    # Помесячная выручка CRM по проектам (в тыс. руб.)
    df_full = df_full.copy()
    if COL_MONTH in df_full.columns:
        _ms = df_full[COL_MONTH].astype(str).str.strip()
        df_full['_month_num'] = _ms.str.split('.').str[0].str.zfill(2)
        # Год: формат 'MM.YY' или 'MM.YYYY'
        _yr = pd.to_numeric(_ms.str.split('.').str[1], errors='coerce').fillna(0).astype(int)
        df_full['_year_num'] = _yr.apply(lambda y: (2000 + y) if (0 < y < 100) else y)
    else:
        df_full['_month_num'] = '00'
        df_full['_year_num'] = 0

    # Строим таблицу
    rows = []
    row_totals = {}  # для подсчётов блоков

    for row_name, crm_projects in ACCOUNTING_ROWS:
        month_values = {}
        total = 0.0

        for m in months:
            val = 0.0

            # 1. CRM-часть (только нужный год, чтобы не суммировать несколько лет)
            if crm_projects:
                mask = (
                    df_full[COL_PROJECT].isin(crm_projects) &
                    (df_full['_month_num'] == m) &
                    (df_full['_year_num'] == analysis_year)
                )
                val += df_full.loc[mask, revenue_col].sum() / 1000  # в тыс. руб.

            # 2. Внешняя часть (из JSON)
            if row_name in ext and isinstance(ext[row_name], dict):
                ext_val = ext[row_name].get(m, 0) or 0
                val += ext_val / 1000  # из руб. в тыс.

            month_values[m] = round(val, 2)
            total += val

        rows.append({'Показатель': row_name, **month_values, 'Total': round(total, 2)})
        row_totals[row_name] = month_values

    # Добавляем промежуточные итоги (Total Ads / Events / Прочие)
    for block_name, row_names in ACCOUNTING_BLOCKS.items():
        block_monthly = {m: 0.0 for m in months}
        for rn in row_names:
            if rn in row_totals:
                for m in months:
                    block_monthly[m] += row_totals[rn].get(m, 0)
        block_total = sum(block_monthly.values())
        rows.append({
            'Показатель': block_name,
            **{m: round(block_monthly[m], 2) for m in months},
            'Total': round(block_total, 2),
        })
        row_totals[block_name] = block_monthly

    # Total Commercial Sales = Ads + Events + Прочие
    commercial_monthly = {m: 0.0 for m in months}
    for block in ['Total Ads Sales', 'Total Events Sales', 'Total Прочие доходы']:
        if block in row_totals:
            for m in months:
                commercial_monthly[m] += row_totals[block].get(m, 0)
    commercial_total = sum(commercial_monthly.values())
    rows.append({
        'Показатель': 'Total Commercial Sales',
        **{m: round(commercial_monthly[m], 2) for m in months},
        'Total': round(commercial_total, 2),
    })
    row_totals['Total Commercial Sales'] = commercial_monthly

    # Разделитель
    rows.append({'Показатель': '── Вычеты из бюджета ──', **{m: '' for m in months}, 'Total': ''})

    # Вычеты (тоже из JSON)
    deduction_rows = ['Медийный бартер ФОНТАНКА', 'Медийный бартер ЭВЕНТЫ',
                      'Взаимозачет/Затраты', 'ФФ/АМ взаимозачет',
                      'Программатик ФОНТАНКА', 'Программатик ФОНТАНКА ТГ',
                      'Программатик ДОКТОР', 'Программатик ДОКТОР ТГ',
                      'Корректировка скидки (комиссия ХШМ)', 'ИРИ/АНО/Петроцентр']
    # Выручка 47 (закупка) — тоже вычет
    deduction_rows.append('Выручка 47 (закупка)')

    # Считаем итог в бюджет отдела
    budget_monthly = dict(commercial_monthly)
    for rn in deduction_rows:
        if rn in row_totals:
            for m in months:
                budget_monthly[m] -= row_totals[rn].get(m, 0)

    budget_total = sum(budget_monthly.values())
    rows.append({
        'Показатель': 'Итого в бюджет отдела рекламы',
        **{m: round(budget_monthly[m], 2) for m in months},
        'Total': round(budget_total, 2),
    })

    # План группы
    plan_monthly = ext.get('_план_группы_по_месяцам_руб', {})
    if plan_monthly:
        plan_row = {m: round(plan_monthly.get(m, 0) / 1000, 2) for m in months}
        plan_total = sum(plan_row.values())
        rows.append({
            'Показатель': 'План группы',
            **plan_row,
            'Total': round(plan_total, 2),
        })

        # Выполнение группы, %
        perf_row = {}
        for m in months:
            p = plan_row.get(m, 0)
            f = budget_monthly.get(m, 0)
            perf_row[m] = round(f / p * 100, 1) if p else None
        perf_total = round(budget_total / plan_total * 100, 1) if plan_total else None
        rows.append({
            'Показатель': 'Выполнение группы, %',
            **perf_row,
            'Total': perf_total,
        })

    # ── KPI по направлениям ──
    kpi_plans = ext.get('_план_KPI_по_месяцам_руб', {})
    if kpi_plans and any(
        any(v for v in kpi_plans.get(d, {}).values() if isinstance(v, (int, float)) and v > 0)
        for d in KPI_DIRECTIONS
    ):
        rows.append({'Показатель': '── KPI ──', **{m: '' for m in months}, 'Total': ''})
        for direction, row_names in KPI_DIRECTIONS.items():
            # Факт по направлению = сумма row_totals
            fact_monthly = {m: 0.0 for m in months}
            for rn in row_names:
                if rn in row_totals:
                    for m in months:
                        fact_monthly[m] += row_totals[rn].get(m, 0)

            plan_dir = kpi_plans.get(direction, {})
            kpi_row = {}
            for m in months:
                p = (plan_dir.get(m, 0) or 0) / 1000
                f = fact_monthly[m]
                kpi_row[m] = round(f / p * 100, 1) if p else None

            plan_total_dir = sum((plan_dir.get(m, 0) or 0) / 1000 for m in months)
            fact_total_dir = sum(fact_monthly.values())
            kpi_total = round(fact_total_dir / plan_total_dir * 100, 1) if plan_total_dir else None

            rows.append({
                'Показатель': f'KPI {direction}, %',
                **kpi_row,
                'Total': kpi_total,
            })


    # Собираем DataFrame
    columns = ['Показатель'] + [month_labels[m] for m in months] + ['Total']
    df_out = pd.DataFrame([
        {'Показатель': r['Показатель'], **{month_labels[m]: r.get(m, '') for m in months}, 'Total': r.get('Total', '')}
        for r in rows
    ], columns=columns)

    return df_out


def build_signals(df_full, revenue_col, monthly_stats, manager_stats, client_stats,
                  rfm_all, mask_no_events, log=print):
    """
    Автоматически находит аномалии и сигналы в данных.
    Возвращает DataFrame с колонками: Приоритет, Категория, Сигнал, Детали.
    """
    signals = []

    def add(priority, category, signal, details=''):
        signals.append({'Приоритет': priority, 'Категория': category,
                        'Сигнал': signal, 'Детали': details})

    # ── 1. Аномалии по месяцам ───────────────────────────────
    if monthly_stats is not None and len(monthly_stats) >= 3:
        rev = monthly_stats['Сумма выручки, тыс. руб.']
        mean_rev = rev.mean()
        std_rev  = rev.std()
        periods  = monthly_stats['Период'].tolist()

        for i, (period, val) in enumerate(zip(periods, rev)):
            if std_rev > 0 and abs(val - mean_rev) > 2 * std_rev:
                direction = '🔺 аномально высокая' if val > mean_rev else '🔻 аномально низкая'
                add('🔴 Высокий', 'Месячная выручка',
                    f'{period}: {direction} выручка',
                    f'{val:,.0f} тыс. vs среднее {mean_rev:,.0f} тыс. (±{std_rev:,.0f})')

        # Падение последних двух месяцев подряд
        if len(rev) >= 3:
            last3 = rev.iloc[-3:].tolist()
            if last3[2] < last3[1] < last3[0]:
                add('🟡 Средний', 'Тренд',
                    f'Выручка падает три месяца подряд',
                    f'{periods[-3]}: {last3[0]:,.0f} → {periods[-2]}: {last3[1]:,.0f} → {periods[-1]}: {last3[2]:,.0f} тыс.')

        # Лучший и худший месяц
        idx_max = rev.idxmax()
        idx_min = rev.idxmin()
        add('🟢 Инфо', 'Месячная выручка',
            f'Лучший месяц: {monthly_stats.loc[idx_max, "Период"]}',
            f'{rev[idx_max]:,.0f} тыс. руб.')
        add('🟢 Инфо', 'Месячная выручка',
            f'Худший месяц: {monthly_stats.loc[idx_min, "Период"]}',
            f'{rev[idx_min]:,.0f} тыс. руб.')

    # ── 2. Концентрация выручки (топ-клиенты) ────────────────
    if client_stats is not None and len(client_stats) > 0:
        total_rev = client_stats['Сумма выручки, тыс. руб.'].sum()
        if total_rev > 0:
            top1_share = client_stats.iloc[0]['Сумма выручки, тыс. руб.'] / total_rev * 100
            top5_share = client_stats.head(5)['Сумма выручки, тыс. руб.'].sum() / total_rev * 100
            top1_name  = client_stats.iloc[0]['КОНЕЧНЫЙ_КЛИЕНТ']

            if top1_share > 20:
                add('🔴 Высокий', 'Концентрация',
                    f'Один клиент даёт {top1_share:.1f}% выручки — высокий риск',
                    f'{top1_name}: {client_stats.iloc[0]["Сумма выручки, тыс. руб."]:,.0f} тыс.')
            elif top1_share > 10:
                add('🟡 Средний', 'Концентрация',
                    f'Топ-клиент даёт {top1_share:.1f}% выручки',
                    f'{top1_name}: {client_stats.iloc[0]["Сумма выручки, тыс. руб."]:,.0f} тыс.')

            if top5_share > 60:
                add('🟡 Средний', 'Концентрация',
                    f'Топ-5 клиентов дают {top5_share:.1f}% выручки',
                    ', '.join(client_stats.head(5)['КОНЕЧНЫЙ_КЛИЕНТ'].tolist()))

    # ── 3. Клиенты с одним заказом (риск оттока) ─────────────
    if 'КОНЕЧНЫЙ_КЛИЕНТ' in df_full.columns:
        order_counts = df_full.groupby('КОНЕЧНЫЙ_КЛИЕНТ')[revenue_col].agg(['count', 'sum'])
        single_order = order_counts[order_counts['count'] == 1]
        single_rev_share = single_order['sum'].sum() / order_counts['sum'].sum() * 100 if order_counts['sum'].sum() > 0 else 0
        add('🟡 Средний' if single_rev_share > 30 else '🟢 Инфо', 'Лояльность',
            f'{len(single_order)} клиентов с единственным заказом ({single_rev_share:.1f}% выручки)',
            f'Потенциал повторных продаж: {single_order["sum"].sum()/1000:,.0f} тыс. руб.')

    # ── 4. RFM-сигналы ────────────────────────────────────────
    if rfm_all is not None and len(rfm_all) > 0:
        at_risk = rfm_all[rfm_all['Segment'] == 'At Risk']
        hibernating = rfm_all[rfm_all['Segment'] == 'Hibernating']

        if len(at_risk) > 0:
            at_risk_rev = at_risk['Monetary_thousands'].sum()
            add('🔴 Высокий', 'RFM / Отток',
                f'{len(at_risk)} клиентов "At Risk" — давно не покупали',
                f'Суммарная ценность: {at_risk_rev:,.0f} тыс. руб. | Топ: ' +
                ', '.join(at_risk.head(3)['КОНЕЧНЫЙ_КЛИЕНТ'].tolist()))

        if len(hibernating) > 0:
            hib_rev = hibernating['Monetary_thousands'].sum()
            add('🟡 Средний', 'RFM / Отток',
                f'{len(hibernating)} клиентов "Hibernating" — неактивны',
                f'Суммарная ценность: {hib_rev:,.0f} тыс. руб.')

        champions = rfm_all[rfm_all['Segment'] == 'Champions']
        if len(champions) > 0:
            add('🟢 Инфо', 'RFM / Рост',
                f'{len(champions)} клиентов-чемпионов — активно покупают',
                ', '.join(champions.head(5)['КОНЕЧНЫЙ_КЛИЕНТ'].tolist()))

    # ── 5. Менеджеры ─────────────────────────────────────────
    if manager_stats is not None and len(manager_stats) >= 2:
        rev_m = manager_stats['Сумма выручки, тыс. руб.']
        mean_m = rev_m.mean()
        top_mgr = manager_stats.iloc[0]
        bot_mgr = manager_stats.iloc[-1]
        gap = top_mgr['Сумма выручки, тыс. руб.'] / bot_mgr['Сумма выручки, тыс. руб.'] if bot_mgr['Сумма выручки, тыс. руб.'] > 0 else 0

        if gap > 5:
            add('🟡 Средний', 'Менеджеры',
                f'Разрыв между топ и аутсайдером в {gap:.1f}x',
                f'{top_mgr[COL_MANAGER]}: {top_mgr["Сумма выручки, тыс. руб."]:,.0f} тыс. vs '
                f'{bot_mgr[COL_MANAGER]}: {bot_mgr["Сумма выручки, тыс. руб."]:,.0f} тыс.')

        add('🟢 Инфо', 'Менеджеры',
            f'Лучший менеджер: {top_mgr[COL_MANAGER]}',
            f'{top_mgr["Сумма выручки, тыс. руб."]:,.0f} тыс. руб.')

    # ── 6. Средний чек ────────────────────────────────────────
    rev_no_events = df_full.loc[mask_no_events, revenue_col].dropna()
    if len(rev_no_events) > 0:
        avg_check = rev_no_events.mean() / 1000
        median_check = rev_no_events.median() / 1000
        ratio = avg_check / median_check if median_check > 0 else 1

        if ratio > 3:
            add('🟡 Средний', 'Чеки',
                f'Средний чек в {ratio:.1f}x выше медианного — влияют крупные сделки',
                f'Среднее: {avg_check:,.1f} тыс. | Медиана: {median_check:,.1f} тыс.')

    # ── 7. Нулевые и аномальные суммы ────────────────────────
    zero_rev = (df_full[revenue_col].fillna(0) == 0).sum()
    if zero_rev > 0:
        add('🟡 Средний', 'Качество данных',
            f'{zero_rev} заказов с нулевой выручкой',
            'Проверь: возможно незакрытые сделки или технические строки')

    if not signals:
        add('🟢 Инфо', 'Общее', 'Явных аномалий не обнаружено', 'Данные выглядят стабильно')

    df_signals = pd.DataFrame(signals)
    priority_order = {'🔴 Высокий': 0, '🟡 Средний': 1, '🟢 Инфо': 2}
    df_signals['_sort'] = df_signals['Приоритет'].map(priority_order).fillna(3)
    df_signals = df_signals.sort_values('_sort').drop(columns='_sort').reset_index(drop=True)
    return df_signals


def style_workbook(wb, log=print):
    """
    Применяет единое оформление ко всем листам книги:
    - оранжевые заголовки с белым текстом
    - чередующиеся строки (зебра)
    - красный фон для отрицательных чисел
    - жирный шрифт для итоговых строк
    - заморозка первой строки
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ORANGE      = PatternFill('solid', fgColor='F38120')
    ROW_ALT     = PatternFill('solid', fgColor='FFF4EC')
    ROW_NORMAL  = PatternFill('solid', fgColor='FFFFFF')
    RED_FILL    = PatternFill('solid', fgColor='FDECEA')
    TOTAL_FILL  = PatternFill('solid', fgColor='FDE8D0')
    SIGNAL_RED  = PatternFill('solid', fgColor='FDECEA')
    SIGNAL_YEL  = PatternFill('solid', fgColor='FEF9E7')
    SIGNAL_GRN  = PatternFill('solid', fgColor='EAFAF1')

    HDR_FONT    = Font(name='Segoe UI', bold=True, color='FFFFFF', size=10)
    BODY_FONT   = Font(name='Segoe UI', size=9)
    TOTAL_FONT  = Font(name='Segoe UI', bold=True, size=9)
    THIN        = Side(style='thin', color='E8E8E8')
    BORDER      = Border(bottom=THIN)

    TOTAL_KEYWORDS = {'итого', 'total', 'план', 'выполнение', 'grand', '═══', '──'}
    SEPARATOR_CHARS = {'═', '─', '—'}

    try:
        for ws in wb.worksheets:
            if ws.title in ('📊 Графики',):
                continue

            max_row = ws.max_row
            max_col = ws.max_column
            if max_row < 2:
                continue

            # Заголовок
            for col in range(1, max_col + 1):
                cell = ws.cell(1, col)
                cell.fill = ORANGE
                cell.font = HDR_FONT
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = BORDER
            ws.row_dimensions[1].height = 22
            ws.freeze_panes = 'A2'

            # Тело
            is_signals = ws.title.startswith('00_Сигналы')

            for row in range(2, max_row + 1):
                first_val = str(ws.cell(row, 1).value or '').strip()

                # Строка-разделитель (═══ или ──)
                is_separator = any(c in first_val for c in SEPARATOR_CHARS)
                # Итоговая строка
                is_total = any(k in first_val.lower() for k in TOTAL_KEYWORDS)

                # Фон строки
                if is_separator:
                    row_fill = PatternFill('solid', fgColor='F0F0F0')
                    row_font = Font(name='Segoe UI', bold=True, color='888888', size=8)
                elif is_total:
                    row_fill = TOTAL_FILL
                    row_font = TOTAL_FONT
                elif is_signals:
                    priority_val = str(ws.cell(row, 1).value or '')
                    if '🔴' in priority_val:
                        row_fill = SIGNAL_RED
                    elif '🟡' in priority_val:
                        row_fill = SIGNAL_YEL
                    else:
                        row_fill = SIGNAL_GRN
                    row_font = BODY_FONT
                else:
                    row_fill = ROW_ALT if row % 2 == 0 else ROW_NORMAL
                    row_font = BODY_FONT

                for col in range(1, max_col + 1):
                    cell = ws.cell(row, col)
                    cell.fill = row_fill
                    cell.font = row_font
                    cell.border = BORDER
                    cell.alignment = Alignment(vertical='center')

                    # Отрицательные числа — красный фон
                    if not is_separator and not is_total:
                        try:
                            if isinstance(cell.value, (int, float)) and cell.value < 0:
                                cell.fill = RED_FILL
                                cell.font = Font(name='Segoe UI', size=9, color='C0392B')
                        except Exception:
                            pass

                ws.row_dimensions[row].height = 18

        log("Оформление листов применено ✅")
    except Exception as e:
        log(f"⚠ Оформление не применено: {e}")


def run_data_quality_checks(df, log=print):
    report = {}
    critical_errors = []

    required_columns = [COL_ORDER, COL_DATE, COL_REVENUE]
    missing_cols = [c for c in required_columns if c not in df.columns]

    report['Всего строк'] = len(df)
    report['Отсутствующие обязательные колонки'] = ", ".join(missing_cols) if missing_cols else "Нет"

    if missing_cols:
        critical_errors.append(f"Отсутствуют обязательные колонки: {missing_cols}")

    if COL_ORDER in df.columns:
        report['Дубликатов заказов'] = df.duplicated(subset=[COL_ORDER]).sum()
    else:
        report['Дубликатов заказов'] = "Колонка отсутствует"

    if COL_REVENUE in df.columns:
        report['Отрицательная выручка'] = (df[COL_REVENUE] < 0).sum()
        report['Нулевая выручка']       = (df[COL_REVENUE] == 0).sum()
        report['Пропущенная выручка']   = df[COL_REVENUE].isna().sum()
        if report['Отрицательная выручка'] > 0:
            critical_errors.append("Обнаружена отрицательная выручка")

    if COL_DATE in df.columns:
        temp_dates = pd.to_datetime(df[COL_DATE], errors='coerce')
        report['Некорректные/пропущенные даты'] = temp_dates.isna().sum()
        report['Будущие даты'] = (temp_dates > pd.Timestamp.today()).sum()
        if report['Будущие даты'] > 0:
            log(f"⚠ Предупреждение: {report['Будущие даты']} дат в будущем (не критично)")

    if any(c in df.columns for c in [COL_CLIENT, COL_CLIENT_RA, COL_REKLAMD]):
        report['Пустые поля клиентов (суммарно)'] = sum(
            (df[col].astype(str).str.strip() == '').sum()
            for col in [COL_CLIENT_RA, COL_CLIENT, COL_REKLAMD]
            if col in df.columns
        )

    if COL_REVENUE in df.columns:
        try:
            q99 = df[COL_REVENUE].quantile(0.99)
            report['Аномально большие чеки'] = (df[COL_REVENUE] > q99 * 3).sum()
        except Exception:
            report['Аномально большие чеки'] = "Ошибка расчета"

    quality_df = pd.DataFrame(list(report.items()), columns=['Метрика', 'Значение'])
    return quality_df, critical_errors


# ============================================================
# ГЛАВНАЯ ФУНКЦИЯ — запускается из app.py
# ============================================================

def run_analytics(input_path: str, output_path: str, log=print,
                  manager_plan: dict = None, date_by: str = 'order'):
    """
    Запускает полный цикл аналитики.
    input_path   — путь к входному .xlsx
    output_path  — путь к выходному .xlsx
    log          — функция для вывода сообщений (print или GUI-лог)
    manager_plan — dict {имя_менеджера: план_руб} или None
    date_by      — 'order' (дата заказа, по умолчанию) или 'payment' (дата оплаты)

    Возвращает: dict с ключами:
        'output_path'       — путь к созданному файлу
        'crm_total'         — итог CRM, тыс. руб.
        'crm_paydate_total' — итог CRM по дате оплаты, тыс. руб.
        'external_total'    — внешние доходы, тыс. руб.
        'grand_total'       — CRM + внешние, тыс. руб. (по выбранной дате)
        'verified_total'    — верифицированная цель, тыс. руб.
        'deviation_pct'     — процент отклонения от верифицированной цели
    """

    if not os.path.exists(input_path):
        raise FileNotFoundError(f"Файл не найден: {input_path}")

    # ── 0. Загрузка verified_figures.json (или fallback на константы) ──
    verified = None
    vf_candidates = [
        os.path.join(os.path.dirname(input_path), VERIFIED_FIGURES_JSON),
        VERIFIED_FIGURES_JSON,
        os.path.join(os.path.dirname(os.path.abspath(__file__)), VERIFIED_FIGURES_JSON),
    ]
    for vf_path in vf_candidates:
        if os.path.exists(vf_path):
            verified = load_verified_figures(vf_path)
            if verified:
                log(f"Верифицированные цифры подгружены из {os.path.basename(vf_path)}")
                break

    if verified:
        vf_total_with_prog = verified.get('total_with_prog', VERIFIED_TOTAL_WITH_PROG)
        vf_total_no_prog   = verified.get('total_with_barter_no_prog', VERIFIED_TOTAL_WITH_BARTER_NO_PROG)
        vf_ads_no_events   = verified.get('advertising_no_events', VERIFIED_ADVERTISING_NO_EVENTS)
        ext_prog_total     = verified.get('programmatic_external', EXTERNAL_PROGRAMMATIC_TOTAL)
        ext_other_total    = verified.get('other_external_income', EXTERNAL_OTHER_INCOME)
    else:
        log("⚠ verified_figures.json не найден — использую константы из analytics.py")
        vf_total_with_prog = VERIFIED_TOTAL_WITH_PROG
        vf_total_no_prog   = VERIFIED_TOTAL_WITH_BARTER_NO_PROG
        vf_ads_no_events   = VERIFIED_ADVERTISING_NO_EVENTS
        ext_prog_total     = EXTERNAL_PROGRAMMATIC_TOTAL
        ext_other_total    = EXTERNAL_OTHER_INCOME

    # ── 1. Загрузка ──────────────────────────────────────────
    log("Загрузка файла...")
    df_raw = pd.read_excel(input_path, dtype=object)
    df_raw.columns = df_raw.columns.map(lambda x: x.strip() if isinstance(x, str) else x)
    log(f"Загружено строк: {len(df_raw)}")

    # ── 2. Базовая очистка ───────────────────────────────────
    if COL_MONTH in df_raw.columns:
        df_raw = df_raw.loc[
            ~df_raw[COL_MONTH].astype(str).str.strip().str.lower().eq("итого")
        ]
    df_raw = df_raw.dropna(how="all").reset_index(drop=True)

    # ── 3. Парсинг денег ─────────────────────────────────────
    df_raw[COL_REVENUE] = df_raw[COL_REVENUE].apply(parse_money)

    # ── 4. Контроль качества ─────────────────────────────────
    log("Проверка качества данных...")
    quality_report, critical_errors = run_data_quality_checks(df_raw, log=log)

    if critical_errors:
        error_text = "\n".join(f" - {e}" for e in critical_errors)
        raise ValueError(f"Критические ошибки в данных:\n{error_text}")

    log("Качество данных: ОК")

    # ── 5. Основной датасет ──────────────────────────────────
    df_full = df_raw.copy()
    revenue_col = COL_REVENUE

    # ── 6. Клиенты и нормализация ────────────────────────────
    df_full['КОНЕЧНЫЙ_КЛИЕНТ'] = df_full.apply(pick_client, axis=1)
    df_full['КОНЕЧНЫЙ_КЛИЕНТ'] = df_full['КОНЕЧНЫЙ_КЛИЕНТ'].apply(normalize_client)

    df_full['IS_EVENT_CLIENT'] = (
        df_full['КОНЕЧНЫЙ_КЛИЕНТ'].str.upper()
        .str.contains('ФЕСТ|ФЕСТИВАЛ|МЕРОПРИЯТИ', na=False)
    )

    # ── 7. Отрасли ───────────────────────────────────────────
    df_full['ОТРАСЛЬ_КЛИЕНТА'] = df_full.apply(pick_industry, axis=1)
    df_full['ОТРАСЛЬ_КЛИЕНТА_НОРМ'] = df_full.apply(classify_industry, axis=1)

    # ── 7.5. Парсинг даты оплаты ─────────────────────────────
    # Колонка "ДатаСуммаОплаты_" имеет формат "25.02.2025 135000"
    # Используется как опциональный источник месячной разбивки (бьётся с бухгалтерией)
    if 'ДатаСуммаОплаты_' in df_full.columns:
        import re as _re
        _date_pat = _re.compile(r'(\d{2}\.\d{2}\.\d{4})')
        def _parse_paydate(s):
            if pd.isna(s):
                return None
            m = _date_pat.match(str(s).strip())
            if m:
                try:
                    return pd.to_datetime(m.group(1), format='%d.%m.%Y')
                except Exception:
                    return None
            return None
        df_full['Дата_оплаты'] = df_full['ДатаСуммаОплаты_'].apply(_parse_paydate)
    else:
        df_full['Дата_оплаты'] = pd.NaT

    # ── 8. Категории выручки и скидок ────────────────────────
    df_full['Категория выручки'] = df_full[revenue_col].apply(categorize_revenue_amount)

    if COL_DISCOUNT_PCT in df_full.columns:
        df_full['Скидка_%_число'] = pd.to_numeric(
            df_full[COL_DISCOUNT_PCT].astype(str)
            .str.replace(',', '.', regex=False)
            .str.replace('%', '', regex=False)
            .str.strip(),
            errors='coerce'
        )
        df_full['Категория_скидки'] = df_full['Скидка_%_число'].apply(bucket_discount)

    # ── 9. Маски фильтрации ──────────────────────────────────
    if COL_PROJECT in df_full.columns:
        mask_no_events = ~df_full[COL_PROJECT].fillna('').isin(EXCLUDE_PROJECTS)
    else:
        mask_no_events = pd.Series(True, index=df_full.index)

    if COL_BARTER in df_full.columns:
        mask_no_barter = (
            df_full[COL_BARTER].astype(str).str.strip().str.lower() != 'да'
        )
    else:
        mask_no_barter = pd.Series(True, index=df_full.index)

    if COL_PROJECT in df_full.columns:
        mask_no_prog = ~df_full[COL_PROJECT].fillna('').str.contains(
            'программатик|programmatic', case=False, regex=True
        )
        if PROGRAMMATIC_PROJECTS:
            mask_no_prog = mask_no_prog & ~df_full[COL_PROJECT].fillna('').isin(PROGRAMMATIC_PROJECTS)
    else:
        mask_no_prog = pd.Series(True, index=df_full.index)

    # ── 9.5. Загрузка внешних доходов (для agregates по периодам) ──
    # Эти цифры добавляются к CRM-выручке в month/quarter/season статистике,
    # чтобы итоги совпадали с бухгалтерией. Клиенты/менеджеры/отрасли не трогаем.
    ext_monthly = None
    _ext_json_found = None
    json_candidates = [
        os.path.join(os.path.dirname(input_path), EXTERNAL_INCOME_JSON),
        EXTERNAL_INCOME_JSON,
        os.path.join(os.path.dirname(os.path.abspath(__file__)), EXTERNAL_INCOME_JSON),
    ]
    for json_path in json_candidates:
        if os.path.exists(json_path):
            ext_monthly = get_external_monthly_totals(json_path)
            _ext_json_found = json_path
            if ext_monthly:
                log(f"Внешние доходы подгружены: {sum(ext_monthly.values())/1000:,.0f} тыс. руб.")
            break

    validate_analytics_consistency(verified, _ext_json_found, log=log)

    # Полный итог всех внешних статей (для сверки с бухгалтерией)
    full_external_k = get_full_external_total(_ext_json_found) / 1000 if _ext_json_found else 0.0

    # ── 10. Месячная статистика ──────────────────────────────
    monthly_stats = None
    if COL_MONTH in df_full.columns:
        log("Считаю месячную статистику...")
        df_full['Дата_месяц'] = df_full[COL_MONTH].apply(parse_month)
        grp_key = df_full['Дата_месяц'].dt.strftime('%m.%Y')

        monthly_stats = (
            df_full.dropna(subset=['Дата_месяц'])
            .groupby(grp_key)
            .agg({revenue_col: ['sum', 'count']})
            .round(2)
        )
        monthly_stats.columns = ['Сумма выручки, руб.', 'Количество заказов']

        monthly_mean = (
            df_full.loc[mask_no_events].dropna(subset=['Дата_месяц'])
            .groupby(df_full.loc[mask_no_events, 'Дата_месяц'].dt.strftime('%m.%Y'))[revenue_col]
            .mean().round(2)
        )
        monthly_stats['Средний чек, руб.'] = monthly_mean
        monthly_stats['Сумма выручки, тыс. руб.'] = monthly_stats['Сумма выручки, руб.'] / 1000
        monthly_stats = monthly_stats.reset_index().rename(columns={'Дата_месяц': 'Период'})

        # ── Выручка по дате оплаты (для сверки с бухгалтерией) ──
        if df_full['Дата_оплаты'].notna().any():
            paydate_sums = (
                df_full.dropna(subset=['Дата_оплаты'])
                .groupby(df_full['Дата_оплаты'].dt.strftime('%m.%Y'))[revenue_col]
                .sum() / 1000
            ).round(2).to_dict()
            monthly_stats['Выручка по дате оплаты, тыс. руб.'] = (
                monthly_stats['Период'].map(paydate_sums).fillna(0).round(2)
            )

        # ── YoY (год к году) если в данных > 1 года ──────────
        if 'Дата_месяц' in df_full.columns:
            _years = df_full['Дата_месяц'].dt.year.dropna().unique()
            if len(_years) > 1:
                _prev_year_map = {}
                for _, mrow in monthly_stats.iterrows():
                    try:
                        _pstr = mrow['Период']  # формат 'MM.YYYY'
                        _pm, _py = _pstr.split('.')
                        _prev = f"{_pm}.{int(_py) - 1}"
                        _prev_year_map[_pstr] = _prev
                    except Exception:
                        pass
                _rev_by_period = monthly_stats.set_index('Период')['Сумма выручки, тыс. руб.'].to_dict()
                monthly_stats['Пред. год, тыс. руб.'] = (
                    monthly_stats['Период'].map(lambda p: _rev_by_period.get(_prev_year_map.get(p)))
                )
                monthly_stats['YoY, %'] = monthly_stats.apply(
                    lambda r: round(
                        (r['Сумма выручки, тыс. руб.'] / r['Пред. год, тыс. руб.'] - 1) * 100, 1
                    )
                    if pd.notna(r.get('Пред. год, тыс. руб.')) and r['Пред. год, тыс. руб.'] != 0
                    else None,
                    axis=1
                )

        # Добавляем колонки с внешними доходами, если JSON есть
        if ext_monthly is not None:
            def get_ext(period):
                # period формат '07.2025' → ключ '07'
                try:
                    m = str(period).split('.')[0].zfill(2)
                    return ext_monthly.get(m, 0) / 1000  # в тыс. руб.
                except Exception:
                    return 0

            monthly_stats['Внешние доходы, тыс. руб.'] = (
                monthly_stats['Период'].apply(get_ext).round(2)
            )
            monthly_stats['Итого с внешними, тыс. руб.'] = (
                monthly_stats['Сумма выручки, тыс. руб.'] + monthly_stats['Внешние доходы, тыс. руб.']
            ).round(2)
            # Итог по дате оплаты + внешние (если колонка есть)
            if 'Выручка по дате оплаты, тыс. руб.' in monthly_stats.columns:
                monthly_stats['Итого по оплате + внешние, тыс. руб.'] = (
                    monthly_stats['Выручка по дате оплаты, тыс. руб.'] + monthly_stats['Внешние доходы, тыс. руб.']
                ).round(2)

    # ── 11. Топ клиентов ─────────────────────────────────────
    log("Считаю топ клиентов...")
    client_df_for_rank = df_full.loc[
        ~(df_full[COL_PROJECT].fillna('').isin(EXCLUDE_PROJECTS) | df_full['IS_EVENT_CLIENT'])
    ] if COL_PROJECT in df_full.columns else df_full
    client_stats = _rev_stats(client_df_for_rank, 'КОНЕЧНЫЙ_КЛИЕНТ', revenue_col)

    # ── 12. Топ менеджеров ───────────────────────────────────
    manager_stats = None
    if COL_MANAGER in df_full.columns:
        log("Считаю статистику по менеджерам...")
        manager_df = df_full.loc[~df_full[COL_MANAGER].isin(EXCLUDE_MANAGERS)]
        mgr_no_events = manager_df.index.isin(df_full.index[mask_no_events])
        manager_stats = _rev_stats(manager_df, COL_MANAGER, revenue_col, avg_mask=mgr_no_events)

    # ── 13. По отраслям ──────────────────────────────────────
    log("Считаю по отраслям...")
    industry_stats = _rev_stats(df_full, 'ОТРАСЛЬ_КЛИЕНТА_НОРМ', revenue_col, avg_mask=mask_no_events)

    # ── 13а. Выручка по бизнес-группам (логика бухгалтерии) ──
    log("Считаю выручку по группам...")
    df_full['БИЗНЕС_ГРУППА'] = (
        df_full[COL_PROJECT].map(PROJECT_GROUP_MAP).fillna('НЕ КЛАССИФИЦИРОВАНО')
        if COL_PROJECT in df_full.columns else 'НЕ КЛАССИФИЦИРОВАНО'
    )

    group_stats = (
        df_full.groupby('БИЗНЕС_ГРУППА')
        .agg({revenue_col: ['sum', 'count']}).round(2)
    )
    group_stats.columns = ['Выручка CRM, руб.', 'Заказов']
    group_stats['Выручка CRM, тыс. руб.'] = (group_stats['Выручка CRM, руб.'] / 1000).round(2)
    group_stats = (
        group_stats.reset_index()
        .sort_values('Выручка CRM, тыс. руб.', ascending=False)
    )

    # Сводка по группам — CRM vs бухгалтерия (без попытки сложить, чтобы не дублировать)
    crm_ads      = df_full.loc[df_full['БИЗНЕС_ГРУППА'].isin(['Реклама Фонтанка', 'Реклама Доктор']), revenue_col].sum()
    crm_prog     = df_full.loc[df_full['БИЗНЕС_ГРУППА'] == 'Программатик', revenue_col].sum()
    crm_events   = df_full.loc[df_full['БИЗНЕС_ГРУППА'] == 'Мероприятия', revenue_col].sum()
    crm_47other  = df_full.loc[df_full['БИЗНЕС_ГРУППА'] == '47News / Прочее', revenue_col].sum()
    crm_unclass  = df_full.loc[df_full['БИЗНЕС_ГРУППА'] == 'НЕ КЛАССИФИЦИРОВАНО', revenue_col].sum()
    crm_total    = df_full[revenue_col].sum()
    # rev_reklama нужна уже в group_summary_rows — считаем здесь
    rev_reklama  = df_full.loc[mask_no_prog & mask_no_events, revenue_col].sum() / 1000

    # Факт из external_income.json по категориям
    _ext_cats_gs = get_external_totals_by_category(_ext_json_found) if _ext_json_found else {}
    _gs_prog_k    = _ext_cats_gs.get('programmatic', 0) / 1000
    _gs_barter_k  = _ext_cats_gs.get('barter', 0) / 1000
    _gs_iri_k     = _ext_cats_gs.get('iri_grants', 0) / 1000
    _gs_recsys_k  = _ext_cats_gs.get('recsys', 0) / 1000
    _gs_ecom_k    = _ext_cats_gs.get('ecom', 0) / 1000
    _gs_47plan_k  = _ext_cats_gs.get('47_plan', 0) / 1000
    _gs_deduct_k  = _ext_cats_gs.get('deductions', 0) / 1000
    _gs_income_k  = _ext_cats_gs.get('total_income', 0) / 1000

    # Итоговые расчётные показатели для сверки
    _gs_crm_plus_prog_k  = crm_total / 1000 + _gs_prog_k
    _gs_crm_plus_all_k   = crm_total / 1000 + _gs_income_k
    _gs_crm_no_prog_bar  = crm_total / 1000 - crm_prog / 1000 + _gs_barter_k

    def _gs_dev(fact, target):
        if target and target != 0:
            return f"{round((fact / target - 1) * 100, 2):+.2f}%"
        return '—'

    vf_prog_k = vf_total_with_prog / 1000 if vf_total_with_prog else None
    vf_no_prog_k = vf_total_no_prog / 1000 if vf_total_no_prog else None
    vf_ads_k = vf_ads_no_events / 1000 if vf_ads_no_events else None

    group_summary_rows = [
        ('═══ CRM (расчёт по выгрузке) ═══', ''),
        ('Реклама Фонтанка + Доктор, тыс. руб.',  round(crm_ads / 1000, 2)),
        ('Программатик (CRM-часть), тыс. руб.',   round(crm_prog / 1000, 2)),
        ('Мероприятия, тыс. руб.',                round(crm_events / 1000, 2)),
        ('47News / Прочее, тыс. руб.',            round(crm_47other / 1000, 2)),
        ('НЕ классифицированное, тыс. руб.',      round(crm_unclass / 1000, 2)),
        ('Итого CRM, тыс. руб.',                  round(crm_total / 1000, 2)),
        ('═══ Внешние доходы (external_income.json) ═══', ''),
        ('Программатик (вне CRM), тыс. руб.',     round(_gs_prog_k, 2) if _ext_cats_gs else '—'),
        ('Медийный бартер, тыс. руб.',            round(_gs_barter_k, 2) if _ext_cats_gs else '—'),
        ('ИРИ / Гранты, тыс. руб.',               round(_gs_iri_k, 2) if _ext_cats_gs else '—'),
        ('Рекомендательные системы, тыс. руб.',   round(_gs_recsys_k, 2) if _ext_cats_gs else '—'),
        ('E-com, тыс. руб.',                      round(_gs_ecom_k, 2) if _ext_cats_gs else '—'),
        ('47News (в план), тыс. руб.',            round(_gs_47plan_k, 2) if _ext_cats_gs else '—'),
        ('Вычеты (взаимозачёты/корректировки), тыс. руб.', round(_gs_deduct_k, 2) if _ext_cats_gs else '—'),
        ('Итого внешних доходов (без вычетов), тыс. руб.', round(_gs_income_k, 2) if _ext_cats_gs else '—'),
        ('═══ Сверка с верифицированными цифрами ═══', ''),
        ('CRM + программатик (факт), тыс. руб.',
            round(_gs_crm_plus_prog_k, 2) if _ext_cats_gs else round(crm_total / 1000, 2)),
        ('CRM + все внешние доходы (факт), тыс. руб.',
            round(_gs_crm_plus_all_k, 2) if _ext_cats_gs else round(crm_total / 1000, 2)),
        ('Верифицировано: всего с прогр. (бух.), тыс. руб.',
            round(vf_prog_k, 2) if vf_prog_k else '—'),
        ('Расхождение: всего с прогр., тыс. руб.',
            round(_gs_crm_plus_all_k - vf_prog_k, 2)
            if vf_prog_k and _ext_cats_gs else '—'),
        ('Расхождение: всего с прогр., %',
            _gs_dev(_gs_crm_plus_all_k, vf_prog_k)
            if vf_prog_k and _ext_cats_gs else '—'),
        ('── Без программатика ──', ''),
        ('CRM (без прогр.) + бартер (факт), тыс. руб.',
            round(_gs_crm_no_prog_bar, 2) if _ext_cats_gs else '—'),
        ('Верифицировано: с бартером без прогр. (бух.), тыс. руб.',
            round(vf_no_prog_k, 2) if vf_no_prog_k else '—'),
        ('Расхождение: без программатика, тыс. руб.',
            round(_gs_crm_no_prog_bar - vf_no_prog_k, 2)
            if vf_no_prog_k and _ext_cats_gs else '—'),
        ('Расхождение: без программатика, %',
            _gs_dev(_gs_crm_no_prog_bar, vf_no_prog_k)
            if vf_no_prog_k and _ext_cats_gs else '—'),
        ('── Рекламная без мероприятий ──', ''),
        ('CRM реклама (без мер.), тыс. руб.',     round(rev_reklama, 2)),
        ('Верифицировано: реклама без мер. (бух.), тыс. руб.',
            round(vf_ads_k, 2) if vf_ads_k else '—'),
        ('Расхождение: реклама, тыс. руб.',
            round(rev_reklama - vf_ads_k, 2) if vf_ads_k else '—'),
        ('Расхождение: реклама, %',
            _gs_dev(rev_reklama, vf_ads_k) if vf_ads_k else '—'),
        ('═══ Справочно ═══', ''),
        ('Программатик полный (verified_figures), тыс. руб.',
            round(ext_prog_total / 1000, 2) if ext_prog_total else '—'),
        ('Прочие доходы (verified_figures), тыс. руб.',
            round(ext_other_total / 1000, 2) if ext_other_total else '—'),
    ]
    group_summary_df = pd.DataFrame(group_summary_rows, columns=['Показатель', 'Значение'])

    # Разбираем дату заказа один раз — используется в сезонности и RFM
    if COL_DATE in df_full.columns:
        df_full['Дата_заказа'] = pd.to_datetime(df_full[COL_DATE], errors='coerce')

    # ── 14. Сезонность ───────────────────────────────────────
    seasonal_stats = None
    quarterly_stats = None
    if 'Дата_заказа' in df_full.columns:
        log("Считаю сезонность...")
        df_full['Месяц_число'] = df_full['Дата_заказа'].dt.month
        df_full['Квартал'] = df_full['Дата_заказа'].dt.quarter
        df_full['Сезон'] = df_full['Месяц_число'].apply(
            lambda x: 'Зима' if x in [12,1,2] else
                      'Весна' if x in [3,4,5] else
                      'Лето' if x in [6,7,8] else 'Осень'
        )

        seasonal_stats = (
            df_full.groupby('Сезон')
            .agg({revenue_col: ['sum','count'], 'КОНЕЧНЫЙ_КЛИЕНТ': 'nunique'}).round(2)
        )
        seasonal_stats.columns = ['Выручка, руб.', 'Количество заказов', 'Уникальных клиентов']
        seasonal_stats['Средний чек, руб.'] = (
            df_full.loc[mask_no_events].groupby('Сезон')[revenue_col].mean().round(2)
        )
        seasonal_stats['Выручка, тыс. руб.'] = seasonal_stats['Выручка, руб.'] / 1000
        seasonal_stats = seasonal_stats.reset_index()

        quarterly_stats = (
            df_full.groupby('Квартал')
            .agg({revenue_col: ['sum','count']}).round(2)
        )
        quarterly_stats.columns = ['Выручка, руб.', 'Количество заказов']
        quarterly_stats['Выручка, тыс. руб.'] = quarterly_stats['Выручка, руб.'] / 1000
        avg_q = quarterly_stats['Выручка, тыс. руб.'].mean()
        quarterly_stats['Коэф. сезонности'] = (quarterly_stats['Выручка, тыс. руб.'] / avg_q).round(2)
        quarterly_stats = quarterly_stats.reset_index()

        # ── Добавляем внешние доходы по кварталам и сезонам ──
        if ext_monthly is not None:
            month_to_quarter = {
                '01': 1, '02': 1, '03': 1,
                '04': 2, '05': 2, '06': 2,
                '07': 3, '08': 3, '09': 3,
                '10': 4, '11': 4, '12': 4,
            }
            month_to_season = {
                '12': 'Зима', '01': 'Зима', '02': 'Зима',
                '03': 'Весна', '04': 'Весна', '05': 'Весна',
                '06': 'Лето', '07': 'Лето', '08': 'Лето',
                '09': 'Осень', '10': 'Осень', '11': 'Осень',
            }
            ext_by_quarter = {1: 0, 2: 0, 3: 0, 4: 0}
            ext_by_season = {'Зима': 0, 'Весна': 0, 'Лето': 0, 'Осень': 0}
            for m, v in ext_monthly.items():
                ext_by_quarter[month_to_quarter[m]] += v / 1000
                ext_by_season[month_to_season[m]] += v / 1000

            quarterly_stats['Внешние доходы, тыс. руб.'] = (
                quarterly_stats['Квартал'].map(ext_by_quarter).round(2)
            )
            quarterly_stats['Итого с внешними, тыс. руб.'] = (
                quarterly_stats['Выручка, тыс. руб.'] + quarterly_stats['Внешние доходы, тыс. руб.']
            ).round(2)

            seasonal_stats['Внешние доходы, тыс. руб.'] = (
                seasonal_stats['Сезон'].map(ext_by_season).round(2)
            )
            seasonal_stats['Итого с внешними, тыс. руб.'] = (
                seasonal_stats['Выручка, тыс. руб.'] + seasonal_stats['Внешние доходы, тыс. руб.']
            ).round(2)

    # ── 15. RFM-анализ ───────────────────────────────────────
    rfm_all = rfm_segment_extended = rfm_non_top = None
    if 'Дата_заказа' in df_full.columns and 'КОНЕЧНЫЙ_КЛИЕНТ' in df_full.columns:
        try:
            log("Провожу RFM-анализ...")
            current_date = df_full['Дата_заказа'].max()

            rfm_source = df_full.loc[
                ~(df_full[COL_PROJECT].fillna('').isin(EXCLUDE_PROJECTS) | df_full['IS_EVENT_CLIENT'])
            ] if COL_PROJECT in df_full.columns else df_full

            rfm_data = (
                rfm_source.groupby('КОНЕЧНЫЙ_КЛИЕНТ')
                .agg({'Дата_заказа': 'max', COL_ORDER: 'count', revenue_col: 'sum'})
                .reset_index()
                .rename(columns={
                    'Дата_заказа': 'Последняя_покупка',
                    COL_ORDER: 'Частота',
                    revenue_col: 'Денежная_ценность_руб'
                })
            )
            rfm_data['Recency_days']     = (current_date - rfm_data['Последняя_покупка']).dt.days
            rfm_data['Frequency_orders'] = rfm_data['Частота']
            rfm_data['Monetary_thousands'] = rfm_data['Денежная_ценность_руб'] / 1000
            rfm_data = rfm_data.dropna(subset=['Recency_days','Frequency_orders','Monetary_thousands'])

            q = rfm_data[['Recency_days','Frequency_orders','Monetary_thousands']].quantile([0.25,0.5,0.75])

            def score_r(x):
                return 4 if x <= q['Recency_days'][0.25] else 3 if x <= q['Recency_days'][0.5] else 2 if x <= q['Recency_days'][0.75] else 1

            def score_f(x):
                return 1 if x <= q['Frequency_orders'][0.25] else 2 if x <= q['Frequency_orders'][0.5] else 3 if x <= q['Frequency_orders'][0.75] else 4

            def score_m(x):
                return 1 if x <= q['Monetary_thousands'][0.25] else 2 if x <= q['Monetary_thousands'][0.5] else 3 if x <= q['Monetary_thousands'][0.75] else 4

            rfm_data['R'] = rfm_data['Recency_days'].apply(score_r)
            rfm_data['F'] = rfm_data['Frequency_orders'].apply(score_f)
            rfm_data['M'] = rfm_data['Monetary_thousands'].apply(score_m)
            rfm_data['RFM_Score'] = rfm_data[['R','F','M']].sum(axis=1)

            def segment_customer(score):
                if score >= 10: return 'Champions'
                elif score >= 8: return 'Loyal Customers'
                elif score >= 6: return 'Potential Loyalists'
                elif score >= 4: return 'At Risk'
                else: return 'Hibernating'

            rfm_data['Segment'] = rfm_data['RFM_Score'].apply(segment_customer)
            rfm_all = rfm_data.sort_values(['RFM_Score','Monetary_thousands'], ascending=[False,False]).reset_index(drop=True)

            total_c = rfm_all['КОНЕЧНЫЙ_КЛИЕНТ'].nunique()
            total_r = rfm_all['Monetary_thousands'].sum()

            rfm_segment_extended = (
                rfm_all.groupby('Segment')
                .agg(Клиентов=('КОНЕЧНЫЙ_КЛИЕНТ','count'),
                     Выручка_тыс=('Monetary_thousands','sum'),
                     Средний_RFM=('RFM_Score','mean'))
                .reset_index()
            )
            rfm_segment_extended['Доля клиентов, %'] = (rfm_segment_extended['Клиентов'] / total_c * 100).round(1)
            rfm_segment_extended['Доля выручки, %']  = (rfm_segment_extended['Выручка_тыс'] / total_r * 100).round(1)

            rfm_non_top = rfm_all[~rfm_all['Segment'].isin(['Champions','Loyal Customers'])].copy()

        except Exception as e:
            log(f"RFM-анализ не выполнен: {e}")

    # ── 16. Лояльность клиентов ──────────────────────────────
    log("Считаю лояльность клиентов...")
    client_order_stats = (
        df_full.groupby('КОНЕЧНЫЙ_КЛИЕНТ')
        .agg({COL_ORDER: 'count', revenue_col: 'sum'}).round(2)
    )
    client_order_stats.columns = ['Количество заказов', 'Общая выручка, руб.']
    client_order_stats['Общая выручка, тыс. руб.'] = client_order_stats['Общая выручка, руб.'] / 1000

    order_frequency = client_order_stats['Количество заказов'].value_counts().sort_index()
    order_freq_df = pd.DataFrame({
        'Количество заказов': order_frequency.index,
        'Количество клиентов': order_frequency.values,
        'Доля клиентов, %': (order_frequency.values / order_frequency.sum() * 100).round(1)
    })

    # ── 17. Сводки ───────────────────────────────────────────
    rev_all      = df_full[revenue_col].sum() / 1000
    rev_bez_prog = df_full.loc[mask_no_prog, revenue_col].sum() / 1000
    rev_reklama  = df_full.loc[mask_no_prog & mask_no_events, revenue_col].sum() / 1000

    # Собираем сводку — CRM-расчёт и верифицированные цифры бухгалтерии
    summary_rows = [
        ('— CRM (расчёт по выгрузке) —', ''),
        ('Выручка: все заказы в CRM, тыс. руб.', round(rev_all, 2)),
        ('Выручка: CRM без программатика, тыс. руб.', round(rev_bez_prog, 2)),
        ('Выручка: CRM рекламная без мероприятий, тыс. руб.', round(rev_reklama, 2)),
        ('— Верифицированные (из бухгалтерии) —', ''),
    ]
    if vf_total_with_prog:
        summary_rows.append(('Верифицированная: всего с программатиком, тыс. руб.', round(vf_total_with_prog / 1000, 2)))
    if vf_total_no_prog:
        summary_rows.append(('Верифицированная: с бартером без программатика, тыс. руб.', round(vf_total_no_prog / 1000, 2)))
    if vf_ads_no_events:
        summary_rows.append(('Верифицированная: рекламная без мероприятий, тыс. руб.', round(vf_ads_no_events / 1000, 2)))

    summary_rows.extend([
        ('— Прочее —', ''),
        ('Средний чек БЕЗ мероприятий, тыс. руб.', round(df_full.loc[mask_no_events, revenue_col].mean() / 1000, 2)),
        ('Медианный чек БЕЗ мероприятий, тыс. руб.', round(df_full.loc[mask_no_events, revenue_col].median() / 1000, 2)),
        ('Количество заказов', len(df_full)),
        ('Уникальных клиентов', df_full['КОНЕЧНЫЙ_КЛИЕНТ'].nunique()),
        ('Уникальных менеджеров', df_full[COL_MANAGER].nunique() if COL_MANAGER in df_full.columns else None),
    ])

    summary_df = pd.DataFrame(summary_rows, columns=['Метрика', 'Значение'])

    df_raw_for_compare = df_raw.copy()
    if COL_MONTH in df_raw_for_compare.columns:
        df_raw_for_compare = df_raw_for_compare.loc[
            ~df_raw_for_compare[COL_MONTH].astype(str).str.strip().str.lower()
            .isin(['итого', 'nan', 'none'])
        ]
    df_raw_for_compare[COL_REVENUE] = pd.to_numeric(df_raw_for_compare[COL_REVENUE], errors='coerce')

    # ── Детальная сверка CRM vs бухгалтерия ─────────────────────────
    ext_cats = get_external_totals_by_category(_ext_json_found) if _ext_json_found else {}

    ext_prog_k    = ext_cats.get('programmatic', 0) / 1000
    ext_barter_k  = ext_cats.get('barter', 0) / 1000
    ext_iri_k     = ext_cats.get('iri_grants', 0) / 1000
    ext_recsys_k  = ext_cats.get('recsys', 0) / 1000
    ext_ecom_k    = ext_cats.get('ecom', 0) / 1000
    ext_47plan_k  = ext_cats.get('47_plan', 0) / 1000
    ext_other_k   = ext_cats.get('other', 0) / 1000
    ext_deduct_k  = ext_cats.get('deductions', 0) / 1000
    ext_income_k  = ext_cats.get('total_income', 0) / 1000
    ext_net_k     = ext_cats.get('total_net', 0) / 1000

    # CRM + внешние доходы (без закупочных/технических проводок) — главный показатель для сверки
    crm_plus_ext_income_k = rev_all + ext_income_k
    # CRM + все внешние с учётом вычетов — итог "в бюджет"
    crm_plus_ext_net_k    = rev_all + ext_net_k

    def _dev(fact, target):
        if target and target != 0:
            return round((fact / target - 1) * 100, 2)
        return None

    def _fmt_dev(pct):
        if pct is None:
            return '—'
        sign = '+' if pct >= 0 else ''
        return f"{sign}{pct:.2f}%"

    vf_total_with_prog_k = vf_total_with_prog / 1000 if vf_total_with_prog else None
    vf_total_no_prog_k   = vf_total_no_prog   / 1000 if vf_total_no_prog   else None
    vf_ads_no_events_k   = vf_ads_no_events   / 1000 if vf_ads_no_events   else None

    cmp_rows = []

    # ── Блок 1: данные CRM ──
    cmp_rows += [
        ('═══ CRM (выгрузка) ═══', '', '', ''),
        ('Строк в исходнике',           len(df_raw_for_compare),       '',    ''),
        ('Строк после очистки',         len(df_full),                  '',    ''),
        ('CRM итого, тыс. руб.',        round(rev_all, 2),             '',    ''),
        ('CRM без программатика, тыс.', round(rev_bez_prog, 2),        '',    ''),
        ('CRM рекламная (без мер.), тыс.', round(rev_reklama, 2),      '',    ''),
    ]

    # ── Блок 2: внешние доходы из external_income.json ──
    if ext_cats:
        cmp_rows += [
            ('═══ Внешние доходы (external_income.json) ═══', '', '', ''),
            ('Программатик (все), тыс.',     round(ext_prog_k, 2),   '', ''),
            ('Бартер медийный, тыс.',        round(ext_barter_k, 2), '', ''),
            ('ИРИ / Гранты, тыс.',           round(ext_iri_k, 2),    '', ''),
            ('Рекомендательные системы, тыс.', round(ext_recsys_k, 2), '', ''),
            ('E-com, тыс.',                  round(ext_ecom_k, 2),   '', ''),
            ('47News (в план), тыс.',        round(ext_47plan_k, 2), '', ''),
            ('Прочие доходные, тыс.',        round(ext_other_k, 2),  '', ''),
            ('── Вычеты (закупка/взаимозачёт), тыс.', round(ext_deduct_k, 2), '', ''),
            ('Итого внешние доходы (без вычетов), тыс.', round(ext_income_k, 2), '', ''),
            ('Итого внешние нетто (с вычетами), тыс.',   round(ext_net_k, 2),    '', ''),
        ]

    # ── Блок 3: сверка с верифицированными цифрами ──
    cmp_rows += [
        ('═══ Сверка с бухгалтерией ═══', '', '', ''),
    ]

    # Срез 1: CRM + все внешние доходы vs total_with_prog
    if vf_total_with_prog_k:
        d = _dev(crm_plus_ext_income_k, vf_total_with_prog_k)
        cmp_rows += [
            ('── Срез: всего с программатиком ──', '', '', ''),
            ('CRM + внешние доходы, тыс.',  round(crm_plus_ext_income_k, 2), '', ''),
            ('Верифицировано (бух.), тыс.', round(vf_total_with_prog_k, 2),  '', ''),
            ('Расхождение абс., тыс.',      round(crm_plus_ext_income_k - vf_total_with_prog_k, 2), '', ''),
            ('Расхождение, %',              _fmt_dev(d), '', ''),
        ]

    # Срез 2: CRM без программатика + бартер vs total_with_barter_no_prog
    if vf_total_no_prog_k:
        crm_no_prog_barter_k = rev_bez_prog + ext_barter_k
        d2 = _dev(crm_no_prog_barter_k, vf_total_no_prog_k)
        cmp_rows += [
            ('── Срез: с бартером, без программатика ──', '', '', ''),
            ('CRM (без прогр.) + бартер, тыс.', round(crm_no_prog_barter_k, 2), '', ''),
            ('Верифицировано (бух.), тыс.',      round(vf_total_no_prog_k, 2),   '', ''),
            ('Расхождение абс., тыс.',           round(crm_no_prog_barter_k - vf_total_no_prog_k, 2), '', ''),
            ('Расхождение, %',                   _fmt_dev(d2), '', ''),
        ]

    # Срез 3: CRM рекламная без мероприятий vs advertising_no_events
    if vf_ads_no_events_k:
        d3 = _dev(rev_reklama, vf_ads_no_events_k)
        cmp_rows += [
            ('── Срез: рекламная без мероприятий ──', '', '', ''),
            ('CRM рекламная (без мер.), тыс.', round(rev_reklama, 2),      '', ''),
            ('Верифицировано (бух.), тыс.',    round(vf_ads_no_events_k, 2), '', ''),
            ('Расхождение абс., тыс.',         round(rev_reklama - vf_ads_no_events_k, 2), '', ''),
            ('Расхождение, %',                 _fmt_dev(d3), '', ''),
        ]

    # ── Блок 4: помесячная сверка (если есть monthly_stats) ──
    if monthly_stats is not None and ext_cats:
        cmp_rows += [
            ('═══ Помесячно: CRM + программатик ═══', '', '', ''),
        ]
        for _, mrow in monthly_stats.iterrows():
            period = mrow['Период']
            crm_m  = mrow['Сумма выручки, тыс. руб.']
            ext_m  = mrow.get('Внешние доходы, тыс. руб.', 0) or 0
            total_m = mrow.get('Итого с внешними, тыс. руб.', crm_m + ext_m)
            cmp_rows.append((
                period,
                round(crm_m, 2),
                round(ext_m, 2),
                round(total_m, 2),
            ))

    comparison_metrics = pd.DataFrame(
        cmp_rows,
        columns=['Показатель / Период', 'CRM, тыс. руб.', 'Внешние, тыс. руб.', 'Итого, тыс. руб.']
    )

    # ── 18. Прогноз ──────────────────────────────────────────
    forecast_summary = None
    if monthly_stats is not None and len(monthly_stats) >= 3:
        ts = monthly_stats.set_index('Период')['Сумма выручки, тыс. руб.']
        months = np.arange(len(ts))
        coeff = np.polyfit(months, ts.values, 1)
        forecast_value = coeff[0] * len(months) + coeff[1]
        trend = "↑ растёт" if coeff[0] > 0 else "↓ падает"
        strength = abs(coeff[0]) / ts.values.mean() * 100

        forecast_summary = pd.DataFrame({
            'Метрика': [
                'Прогноз выручки на след. месяц, тыс. руб.',
                'Направление тренда',
                'Сила тренда, % от среднего'
            ],
            'Значение': [
                round(forecast_value, 2),
                trend,
                round(strength, 2)
            ]
        })

    # ── 19а. План по менеджерам ──────────────────────────────
    plan_stats = None
    if manager_plan and manager_stats is not None and len(manager_stats) > 0:
        log("Считаю выполнение плана...")
        plan_df = manager_stats.copy()
        plan_df['План, руб.'] = plan_df[COL_MANAGER].map(manager_plan).fillna(0)
        plan_df['План, тыс. руб.'] = (plan_df['План, руб.'] / 1000).round(2)
        plan_df['Выполнение, %'] = plan_df.apply(
            lambda r: round(r['Сумма выручки, руб.'] / r['План, руб.'] * 100, 1)
            if r['План, руб.'] > 0 else None,
            axis=1
        )
        plan_df['Остаток до плана, тыс. руб.'] = plan_df.apply(
            lambda r: round((r['План, руб.'] - r['Сумма выручки, руб.']) / 1000, 2)
            if r['План, руб.'] > 0 else None,
            axis=1
        )
        plan_df['Статус'] = plan_df['Выполнение, %'].apply(
            lambda x: '✅ Выполнен' if x is not None and x >= 100
            else ('⚠ В работе' if x is not None and x >= 70
            else ('❌ Отстаёт' if x is not None else '— нет плана'))
        )
        plan_stats = plan_df[[
            COL_MANAGER,
            'Сумма выручки, тыс. руб.',
            'План, тыс. руб.',
            'Выполнение, %',
            'Остаток до плана, тыс. руб.',
            'Статус'
        ]].sort_values('Выполнение, %', ascending=False, na_position='last')

    # ── 19б. Топ номенклатур ─────────────────────────────────
    nomen_stats = None
    COL_NOMEN_LOCAL = 'Номенклатура'
    if COL_NOMEN_LOCAL in df_full.columns:
        log("Считаю топ номенклатур...")
        nomen_df = df_full.loc[mask_no_events].copy()
        nomen_stats = (
            nomen_df.groupby(COL_NOMEN_LOCAL)
            .agg(
                Выручка_руб=(revenue_col, 'sum'),
                Заказов=(COL_ORDER, 'count'),
                Клиентов=('КОНЕЧНЫЙ_КЛИЕНТ', 'nunique')
            )
            .reset_index()
            .sort_values('Выручка_руб', ascending=False)
        )
        nomen_stats['Выручка, тыс. руб.'] = (nomen_stats['Выручка_руб'] / 1000).round(2)
        nomen_stats['Средний чек, тыс. руб.'] = (
            nomen_stats['Выручка_руб'] / nomen_stats['Заказов'] / 1000
        ).round(2)
        total_rev = nomen_stats['Выручка_руб'].sum()
        nomen_stats['Доля выручки, %'] = (
            nomen_stats['Выручка_руб'] / total_rev * 100
        ).round(1)
        nomen_stats = nomen_stats[[
            COL_NOMEN_LOCAL,
            'Выручка, тыс. руб.',
            'Доля выручки, %',
            'Заказов',
            'Клиентов',
            'Средний чек, тыс. руб.'
        ]].head(50)

    # ── 19б_2. Анализ скидок ─────────────────────────────────
    discount_stats = None
    if COL_DISCOUNT_PCT in df_full.columns and 'Скидка_%_число' in df_full.columns:
        log("Считаю анализ скидок...")
        df_disc = df_full.loc[
            mask_no_events &
            df_full['Скидка_%_число'].notna() &
            (df_full['Скидка_%_число'] > 0)
        ].copy()

        if len(df_disc) > 0:
            # По клиентам
            disc_by_client = (
                df_disc.groupby('КОНЕЧНЫЙ_КЛИЕНТ')
                .agg(
                    Заказов=(COL_ORDER, 'count'),
                    Выручка_руб=(revenue_col, 'sum'),
                    Ср_скидка=('Скидка_%_число', 'mean'),
                    Макс_скидка=('Скидка_%_число', 'max'),
                )
                .reset_index()
                .sort_values('Выручка_руб', ascending=False)
                .head(30)
            )
            disc_by_client['Выручка, тыс. руб.'] = (disc_by_client['Выручка_руб'] / 1000).round(2)
            disc_by_client['Ср. скидка, %'] = disc_by_client['Ср_скидка'].round(1)
            disc_by_client['Макс. скидка, %'] = disc_by_client['Макс_скидка'].round(1)

            # По менеджерам
            disc_by_mgr = None
            if COL_MANAGER in df_disc.columns:
                disc_by_mgr = (
                    df_disc.groupby(COL_MANAGER)
                    .agg(
                        Заказов=(COL_ORDER, 'count'),
                        Выручка_руб=(revenue_col, 'sum'),
                        Ср_скидка=('Скидка_%_число', 'mean'),
                        Макс_скидка=('Скидка_%_число', 'max'),
                        Доля_со_скидкой=(revenue_col, 'count'),
                    )
                    .reset_index()
                    .sort_values('Ср_скидка', ascending=False)
                )
                disc_by_mgr['Выручка, тыс. руб.'] = (disc_by_mgr['Выручка_руб'] / 1000).round(2)
                disc_by_mgr['Ср. скидка, %'] = disc_by_mgr['Ср_скидка'].round(1)
                disc_by_mgr['Макс. скидка, %'] = disc_by_mgr['Макс_скидка'].round(1)

            # Распределение по бакетам
            disc_bucket_stats = (
                df_disc.groupby('Категория_скидки')
                .agg(
                    Заказов=(COL_ORDER, 'count'),
                    Выручка_руб=(revenue_col, 'sum'),
                )
                .reset_index()
                .sort_values('Выручка_руб', ascending=False)
            )
            disc_bucket_stats['Выручка, тыс. руб.'] = (disc_bucket_stats['Выручка_руб'] / 1000).round(2)
            total_disc_rev = disc_bucket_stats['Выручка_руб'].sum()
            disc_bucket_stats['Доля, %'] = (
                disc_bucket_stats['Выручка_руб'] / total_disc_rev * 100
            ).round(1) if total_disc_rev > 0 else 0

            # Итоговая сводка: общий % заказов со скидкой
            total_orders = len(df_full.loc[mask_no_events])
            orders_with_disc = len(df_disc)
            disc_summary_rows = [
                ('Заказов без мероприятий (всего)', total_orders),
                ('Заказов со скидкой', orders_with_disc),
                ('Доля заказов со скидкой, %', round(orders_with_disc / total_orders * 100, 1) if total_orders else 0),
                ('Средняя скидка по скидочным сделкам, %', round(df_disc['Скидка_%_число'].mean(), 1)),
                ('Медианная скидка, %', round(df_disc['Скидка_%_число'].median(), 1)),
                ('Максимальная скидка, %', round(df_disc['Скидка_%_число'].max(), 1)),
                ('Выручка со скидкой, тыс. руб.', round(df_disc[revenue_col].sum() / 1000, 2)),
                ('Доля выручки со скидкой, %',
                    round(df_disc[revenue_col].sum() /
                          df_full.loc[mask_no_events, revenue_col].sum() * 100, 1)
                    if df_full.loc[mask_no_events, revenue_col].sum() > 0 else 0),
            ]
            disc_summary_df = pd.DataFrame(disc_summary_rows, columns=['Метрика', 'Значение'])

            # Склеиваем в один датафрейм с разделителями
            sep = pd.DataFrame([['═══ Сводка ═══', '']], columns=['Метрика', 'Значение'])
            sep2 = pd.DataFrame([['═══ По бакетам скидок ═══', '']], columns=['Метрика', 'Значение'])
            discount_stats = {
                'summary': disc_summary_df,
                'by_client': disc_by_client[[
                    'КОНЕЧНЫЙ_КЛИЕНТ', 'Выручка, тыс. руб.',
                    'Заказов', 'Ср. скидка, %', 'Макс. скидка, %'
                ]],
                'by_manager': disc_by_mgr[[
                    COL_MANAGER, 'Выручка, тыс. руб.',
                    'Заказов', 'Ср. скидка, %', 'Макс. скидка, %'
                ]] if disc_by_mgr is not None else None,
                'by_bucket': disc_bucket_stats[[
                    'Категория_скидки', 'Заказов', 'Выручка, тыс. руб.', 'Доля, %'
                ]],
            }

    # ── 19в. Сигналы и аномалии ─────────────────────────────
    log("Ищу сигналы и аномалии...")
    signals_df = build_signals(
        df_full, revenue_col, monthly_stats, manager_stats, client_stats,
        rfm_all, mask_no_events, log=log
    )

    # ── 19г. Бухгалтерская таблица ───────────────────────────
    log("Строю бухгалтерскую таблицу...")
    accounting_table = None
    # Ищем JSON рядом со входным файлом, потом в рабочей папке
    json_candidates = [
        os.path.join(os.path.dirname(input_path), EXTERNAL_INCOME_JSON),
        EXTERNAL_INCOME_JSON,
        os.path.join(os.path.dirname(os.path.abspath(__file__)), EXTERNAL_INCOME_JSON),
    ]
    for json_path in json_candidates:
        if os.path.exists(json_path):
            accounting_table = build_accounting_table(df_full, revenue_col, json_path, log=log)
            break
    else:
        log(f"⚠ Файл {EXTERNAL_INCOME_JSON} не найден — бухгалтерская таблица не построена")

    # ── 19. Сборка и экспорт ─────────────────────────────────
    log("Сохраняю отчёт...")
    export_data = {}

    export_data['00_Сигналы'] = signals_df
    if quality_report is not None:
        export_data['00_Data_Quality'] = quality_report
    if monthly_stats is not None:
        export_data['01_Месячная_статистика'] = monthly_stats
    export_data['02_Топ_клиентов'] = client_stats.head(50)
    if manager_stats is not None:
        export_data['03_Топ_менеджеров'] = manager_stats.head(30)
    if plan_stats is not None:
        export_data['03а_План_менеджеров'] = plan_stats
    export_data['04_Отрасли'] = industry_stats
    if nomen_stats is not None:
        export_data['04а_Номенклатура'] = nomen_stats
    if seasonal_stats is not None:
        export_data['05_Сезонность'] = seasonal_stats
    if quarterly_stats is not None:
        export_data['06_Кварталы'] = quarterly_stats
    if rfm_all is not None:
        export_data['07_RFM_все'] = rfm_all
    if rfm_segment_extended is not None:
        export_data['08_RFM_сегменты'] = rfm_segment_extended
    if rfm_non_top is not None:
        export_data['09_RFM_не_топ'] = rfm_non_top
    export_data['10_Распределение_заказов'] = order_freq_df
    export_data['11_Расширенный_сводный']   = summary_df
    export_data['12_Сравнение']             = comparison_metrics
    if forecast_summary is not None:
        export_data['13_Прогноз']           = forecast_summary
    export_data['14_Выручка_по_группам']    = group_stats
    export_data['15_Сводка_по_группам']     = group_summary_df
    if accounting_table is not None:
        export_data['16_Бухгалтерская_таблица'] = accounting_table
    if discount_stats is not None:
        export_data['17_Скидки_сводка']     = discount_stats['summary']
        export_data['17а_Скидки_клиенты']   = discount_stats['by_client']
        if discount_stats['by_manager'] is not None:
            export_data['17б_Скидки_менеджеры'] = discount_stats['by_manager']
        export_data['17в_Скидки_по_бакетам'] = discount_stats['by_bucket']

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for sheet_name, df_sheet in export_data.items():
            safe_name = sheet_name[:31]
            try:
                df_sheet.to_excel(writer, sheet_name=safe_name, index=False)
                ws = writer.sheets[safe_name]
                for col_idx, col in enumerate(df_sheet.columns, 1):
                    try:
                        col_max = df_sheet[col].dropna().astype(str).map(len).max()
                        col_max = int(col_max) if pd.notna(col_max) else 10
                        max_len = max(col_max, len(str(col))) + 2
                    except Exception:
                        max_len = 14
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 60)
            except Exception as e:
                log(f"Ошибка при сохранении листа {safe_name}: {e}")

    # ── Оформление + графики ─────────────────────────────────
    log("Строю графики...")
    try:
        from openpyxl import load_workbook
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference
        from openpyxl.styles import Font

        wb = load_workbook(output_path)
        style_workbook(wb, log=log)

        # Создаём лист с графиками
        ws_charts = wb.create_sheet("📊 Графики", 0)
        ws_charts.sheet_view.showGridLines = False

        # Заголовок листа
        ws_charts['A1'] = "АНАЛИТИКА — ГРАФИКИ"
        ws_charts['A1'].font = Font(name='Segoe UI', size=16, bold=True, color="F38120")
        ws_charts.row_dimensions[1].height = 30

        chart_row = 3  # текущая строка для размещения графиков

        # ── 1. Выручка по месяцам (линейный) ─────────────────
        sheet_monthly = '01_Месячная_статистика'
        if sheet_monthly in wb.sheetnames:
            ws_m = wb[sheet_monthly]
            max_row = ws_m.max_row

            if max_row > 2:
                chart = LineChart()
                chart.title = "Выручка по месяцам, тыс. руб."
                chart.style = 10
                chart.y_axis.title = "тыс. руб."
                chart.x_axis.title = "Период"
                chart.width = 22
                chart.height = 14
                chart.grouping = "standard"
                chart.smooth = True

                rev_col = _find_rev_col(ws_m)
                data = Reference(ws_m, min_col=rev_col, min_row=1, max_row=max_row)
                cats = Reference(ws_m, min_col=1, min_row=2, max_row=max_row)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                chart.series[0].graphicalProperties.line.solidFill = "F38120"
                chart.series[0].graphicalProperties.line.width = 25000
                chart.series[0].smooth = True

                ws_charts.add_chart(chart, f"A{chart_row}")
                chart_row += 25

        # ── 2. Топ-10 менеджеров (горизонтальный столбчатый) ──
        sheet_mgr = '03_Топ_менеджеров'
        if sheet_mgr in wb.sheetnames:
            ws_mgr = wb[sheet_mgr]
            n = min(11, ws_mgr.max_row)

            if n > 2:
                chart = BarChart()
                chart.type = "bar"
                chart.title = "Топ-10 менеджеров, тыс. руб."
                chart.style = 10
                chart.y_axis.title = "Менеджер"
                chart.x_axis.title = "тыс. руб."
                chart.width = 22
                chart.height = 14

                rev_col = _find_rev_col(ws_mgr)
                data = Reference(ws_mgr, min_col=rev_col, min_row=1, max_row=n)
                cats = Reference(ws_mgr, min_col=1, min_row=2, max_row=n)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                chart.series[0].graphicalProperties.solidFill = "F38120"

                ws_charts.add_chart(chart, f"L{chart_row - 25 + 3}")

        # ── 3. Топ-10 клиентов (вертикальный столбчатый) ──────
        sheet_cli = '02_Топ_клиентов'
        if sheet_cli in wb.sheetnames:
            ws_cli = wb[sheet_cli]
            n = min(11, ws_cli.max_row)

            if n > 2:
                chart = BarChart()
                chart.type = "col"
                chart.title = "Топ-10 клиентов, тыс. руб."
                chart.style = 10
                chart.y_axis.title = "тыс. руб."
                chart.width = 22
                chart.height = 14

                rev_col = _find_rev_col(ws_cli)
                data = Reference(ws_cli, min_col=rev_col, min_row=1, max_row=n)
                cats = Reference(ws_cli, min_col=1, min_row=2, max_row=n)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                chart.series[0].graphicalProperties.solidFill = "4A90D9"

                ws_charts.add_chart(chart, f"A{chart_row}")
                chart_row += 25

        # ── 4. Отрасли (круговая диаграмма) ───────────────────
        sheet_ind = '04_Отрасли'
        if sheet_ind in wb.sheetnames:
            ws_ind = wb[sheet_ind]
            n = min(11, ws_ind.max_row)

            if n > 2:
                chart = PieChart()
                chart.title = "Выручка по отраслям"
                chart.style = 10
                chart.width = 18
                chart.height = 14

                rev_col = _find_rev_col(ws_ind)
                data = Reference(ws_ind, min_col=rev_col, min_row=1, max_row=n)
                cats = Reference(ws_ind, min_col=1, min_row=2, max_row=n)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                chart.dataLabels = None

                ws_charts.add_chart(chart, f"L{chart_row - 25 + 3}")

        wb.save(output_path)
        log("Графики добавлены ✅")
    except Exception as e:
        log(f"⚠ Графики не добавлены: {e}")

    log(f"✅ Готово! Отчёт сохранён: {output_path}")

    # ── Сводка расхождения для GUI ───────────────────────────
    crm_total_k       = df_full[revenue_col].sum() / 1000
    crm_paydate_total_k = (
        df_full.dropna(subset=['Дата_оплаты'])[revenue_col].sum() / 1000
        if df_full['Дата_оплаты'].notna().any() else 0
    )
    # external_total_k — только программатик (используется в колонках месячной аналитики)
    external_total_k = sum(ext_monthly.values()) / 1000 if ext_monthly else 0

    # full_external_k — внешние доходы без закупочных/технических статей
    # (теперь get_full_external_total исключает Выручка47-закупка, Взаимозачет/Затраты,
    #  ФФ/АМ взаимозачет, Корректировка скидки — они не реальный доход)
    crm_base_k = crm_paydate_total_k if (date_by == 'payment' and crm_paydate_total_k > 0) else crm_total_k
    grand_total_k = crm_base_k + full_external_k

    verified_total_k = vf_total_with_prog / 1000 if vf_total_with_prog else 0
    deviation_pct = (
        (grand_total_k / verified_total_k - 1) * 100
        if verified_total_k > 0 else None
    )

    report_info = {
        'output_path':           output_path,
        'crm_total':             round(crm_total_k, 2),
        'crm_paydate_total':     round(crm_paydate_total_k, 2),
        'external_total':        round(external_total_k, 2),
        'full_external_total':   round(full_external_k, 2),
        'grand_total':           round(grand_total_k, 2),
        'verified_total':        round(verified_total_k, 2),
        'deviation_pct':         round(deviation_pct, 2) if deviation_pct is not None else None,
        'date_by':               date_by,
    }

    crm_used_k = crm_base_k
    date_label = "по дате оплаты" if (date_by == 'payment' and crm_paydate_total_k > 0) else "по дате заказа"

    log(f"📊 CRM ({date_label}): {crm_used_k:,.0f} тыс. | "
        f"Прогр. (ежемес.): {external_total_k:,.0f} тыс. | "
        f"Все внешние доходы: {full_external_k:,.0f} тыс.")
    if deviation_pct is not None:
        sign = '+' if deviation_pct >= 0 else ''
        verdict = 'отлично' if abs(deviation_pct) < 2 else ('в норме' if abs(deviation_pct) < 5 else 'большое расхождение')
        log(f"📊 Итого CRM + внешние: {grand_total_k:,.0f} тыс. | "
            f"Верифицировано: {verified_total_k:,.0f} тыс. | "
            f"Отклонение: {sign}{deviation_pct:.2f}% — {verdict}")
    else:
        log(f"📊 Итого: {grand_total_k:,.0f} тыс.")

    return report_info
