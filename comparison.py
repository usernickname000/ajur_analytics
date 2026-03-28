# ============================================================
# comparison.py — сравнительный анализ двух периодов
# Вызывается из app.py, использует вспомогательные функции из analytics.py
# ============================================================

import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

from analytics import (
    parse_money, pick_client, normalize_client,
    EXCLUDE_MANAGERS, EXCLUDE_PROJECTS,
    COL_MONTH, COL_ORDER, COL_DATE, COL_MANAGER,
    COL_REVENUE, COL_PROJECT,
    COL_CLIENT, COL_CLIENT_RA, COL_REKLAMD
)


# ============================================================
# ЗАГРУЗКА И ПОДГОТОВКА ОДНОГО ПЕРИОДА
# ============================================================

def _load_period(path: str, log) -> pd.DataFrame:
    """Загружает файл, применяет те же фильтры и нормализацию, что и run_analytics."""
    if not os.path.exists(path):
        raise FileNotFoundError(f"Файл не найден: {path}")

    df = pd.read_excel(path, dtype=object)
    df.columns = df.columns.map(lambda x: x.strip() if isinstance(x, str) else x)

    # Убираем строку "Итого"
    if COL_MONTH in df.columns:
        df = df.loc[~df[COL_MONTH].astype(str).str.strip().str.lower().eq("итого")]
    df = df.dropna(how="all").reset_index(drop=True)

    # Парсим выручку
    df[COL_REVENUE] = df[COL_REVENUE].apply(parse_money)
    df = df.dropna(subset=[COL_REVENUE])
    df = df[df[COL_REVENUE] > 0].reset_index(drop=True)

    # Клиент
    df['КОНЕЧНЫЙ_КЛИЕНТ'] = df.apply(pick_client, axis=1)
    df['КОНЕЧНЫЙ_КЛИЕНТ'] = df['КОНЕЧНЫЙ_КЛИЕНТ'].apply(normalize_client)

    # Исключаем менеджеров вне отдела
    if COL_MANAGER in df.columns:
        df = df.loc[~df[COL_MANAGER].isin(EXCLUDE_MANAGERS)].reset_index(drop=True)

    # Исключаем мероприятия
    if COL_PROJECT in df.columns:
        df = df.loc[~df[COL_PROJECT].fillna('').isin(EXCLUDE_PROJECTS)].reset_index(drop=True)

    log(f"  Загружено строк после фильтров: {len(df)}")
    return df


# ============================================================
# РАСЧЁТЫ ПО ПЕРИОДУ
# ============================================================

def _manager_stats(df: pd.DataFrame) -> pd.DataFrame:
    if COL_MANAGER not in df.columns:
        return pd.DataFrame()
    stats = (
        df.groupby(COL_MANAGER)
        .agg(
            Выручка_руб=(COL_REVENUE, 'sum'),
            Заказов=(COL_ORDER, 'count')
        )
        .reset_index()
        .sort_values('Выручка_руб', ascending=False)
    )
    stats['Выручка_тыс'] = (stats['Выручка_руб'] / 1000).round(2)
    return stats


def _client_stats(df: pd.DataFrame) -> pd.DataFrame:
    stats = (
        df.groupby('КОНЕЧНЫЙ_КЛИЕНТ')
        .agg(
            Выручка_руб=(COL_REVENUE, 'sum'),
            Заказов=(COL_ORDER, 'count')
        )
        .reset_index()
        .sort_values('Выручка_руб', ascending=False)
    )
    stats['Выручка_тыс'] = (stats['Выручка_руб'] / 1000).round(2)
    return stats


# ============================================================
# ФОРМАТИРОВАНИЕ EXCEL
# ============================================================

HEADER_FILL   = PatternFill("solid", start_color="1F3864", end_color="1F3864")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Segoe UI", size=10)
NEW_FILL      = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")   # зелёный
LOST_FILL     = PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")   # красный
POS_FONT      = Font(color="375623", name="Segoe UI", size=9)
NEG_FONT      = Font(color="C00000", name="Segoe UI", size=9)
NEUTRAL_FONT  = Font(name="Segoe UI", size=9)
THIN_BORDER   = Border(
    bottom=Side(style='thin', color="D9D9D9")
)

def _style_header(ws, row_idx: int):
    for cell in ws[row_idx]:
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _auto_width(ws, min_w=10, max_w=50):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, min_w), max_w)

def _write_df(ws, df: pd.DataFrame, start_row=2):
    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = NEUTRAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")


# ============================================================
# ЛИСТ 1: СВОДКА
# ============================================================

def _sheet_summary(wb, df_a, df_b, label_a, label_b):
    ws = wb.create_sheet("01_Сводка")
    ws.freeze_panes = "A2"

    headers = ["Метрика", label_a, label_b, "Δ абс.", "Δ %"]
    ws.append(headers)
    _style_header(ws, 1)

    def pct(a, b):
        if a and a != 0:
            return round((b - a) / abs(a) * 100, 1)
        return None

    metrics = []

    rev_a = round(df_a[COL_REVENUE].sum() / 1000, 2)
    rev_b = round(df_b[COL_REVENUE].sum() / 1000, 2)
    metrics.append(("Выручка, тыс. руб.", rev_a, rev_b))

    ord_a = len(df_a)
    ord_b = len(df_b)
    metrics.append(("Количество заказов", ord_a, ord_b))

    avg_a = round(df_a[COL_REVENUE].mean() / 1000, 2)
    avg_b = round(df_b[COL_REVENUE].mean() / 1000, 2)
    metrics.append(("Средний чек, тыс. руб.", avg_a, avg_b))

    med_a = round(df_a[COL_REVENUE].median() / 1000, 2)
    med_b = round(df_b[COL_REVENUE].median() / 1000, 2)
    metrics.append(("Медианный чек, тыс. руб.", med_a, med_b))

    cli_a = df_a['КОНЕЧНЫЙ_КЛИЕНТ'].nunique()
    cli_b = df_b['КОНЕЧНЫЙ_КЛИЕНТ'].nunique()
    metrics.append(("Уникальных клиентов", cli_a, cli_b))

    if COL_MANAGER in df_a.columns and COL_MANAGER in df_b.columns:
        mgr_a = df_a[COL_MANAGER].nunique()
        mgr_b = df_b[COL_MANAGER].nunique()
        metrics.append(("Активных менеджеров", mgr_a, mgr_b))

    for row_data in metrics:
        name, va, vb = row_data
        delta_abs = round(vb - va, 2) if isinstance(va, (int, float)) and isinstance(vb, (int, float)) else None
        delta_pct = pct(va, vb) if isinstance(va, (int, float)) else None
        row = [name, va, vb, delta_abs, delta_pct]
        ws.append(row)
        r = ws.max_row
        for c in range(1, 6):
            ws.cell(r, c).font = NEUTRAL_FONT
            ws.cell(r, c).border = THIN_BORDER
            ws.cell(r, c).alignment = Alignment(vertical="center")
        # Цвет дельты
        if delta_abs is not None:
            font = POS_FONT if delta_abs >= 0 else NEG_FONT
            ws.cell(r, 4).font = font
            ws.cell(r, 5).font = font

    ws.row_dimensions[1].height = 22
    _auto_width(ws)


# ============================================================
# ЛИСТ 2: МЕНЕДЖЕРЫ
# ============================================================

def _sheet_managers(wb, df_a, df_b, label_a, label_b):
    ws = wb.create_sheet("02_Менеджеры")
    ws.freeze_panes = "A2"

    mgr_a = _manager_stats(df_a).rename(columns={
        'Выручка_тыс': f'Выручка_{label_a}_тыс',
        'Заказов': f'Заказов_{label_a}'
    })
    mgr_b = _manager_stats(df_b).rename(columns={
        'Выручка_тыс': f'Выручка_{label_b}_тыс',
        'Заказов': f'Заказов_{label_b}'
    })

    merged = pd.merge(
        mgr_a[[COL_MANAGER, f'Выручка_{label_a}_тыс', f'Заказов_{label_a}']],
        mgr_b[[COL_MANAGER, f'Выручка_{label_b}_тыс', f'Заказов_{label_b}']],
        on=COL_MANAGER, how='outer'
    ).fillna(0)

    col_a = f'Выручка_{label_a}_тыс'
    col_b = f'Выручка_{label_b}_тыс'
    merged['Δ выручка, тыс.'] = (merged[col_b] - merged[col_a]).round(2)
    merged['Δ %'] = merged.apply(
        lambda r: round((r[col_b] - r[col_a]) / abs(r[col_a]) * 100, 1) if r[col_a] != 0 else None, axis=1
    )
    merged = merged.sort_values(col_b, ascending=False).reset_index(drop=True)

    headers = list(merged.columns)
    ws.append(headers)
    _style_header(ws, 1)
    _write_df(ws, merged)

    # Цвет дельты
    delta_col_idx = headers.index('Δ выручка, тыс.') + 1
    pct_col_idx   = headers.index('Δ %') + 1
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row_idx, delta_col_idx).value
        if val is not None:
            font = POS_FONT if float(val) >= 0 else NEG_FONT
            ws.cell(row_idx, delta_col_idx).font = font
            ws.cell(row_idx, pct_col_idx).font = font

    _auto_width(ws)
    return merged  # вернём для графика


# ============================================================
# ЛИСТ 3: КЛИЕНТЫ
# ============================================================

def _sheet_clients(wb, df_a, df_b, label_a, label_b):
    ws = wb.create_sheet("03_Клиенты")
    ws.freeze_panes = "A2"

    cli_a = _client_stats(df_a).rename(columns={
        'Выручка_тыс': f'Выручка_{label_a}_тыс',
        'Заказов': f'Заказов_{label_a}'
    })
    cli_b = _client_stats(df_b).rename(columns={
        'Выручка_тыс': f'Выручка_{label_b}_тыс',
        'Заказов': f'Заказов_{label_b}'
    })

    merged = pd.merge(
        cli_a[['КОНЕЧНЫЙ_КЛИЕНТ', f'Выручка_{label_a}_тыс', f'Заказов_{label_a}']],
        cli_b[['КОНЕЧНЫЙ_КЛИЕНТ', f'Выручка_{label_b}_тыс', f'Заказов_{label_b}']],
        on='КОНЕЧНЫЙ_КЛИЕНТ', how='outer'
    ).fillna(0)

    col_a = f'Выручка_{label_a}_тыс'
    col_b = f'Выручка_{label_b}_тыс'
    merged['Δ выручка, тыс.'] = (merged[col_b] - merged[col_a]).round(2)
    merged['Δ %'] = merged.apply(
        lambda r: round((r[col_b] - r[col_a]) / abs(r[col_a]) * 100, 1) if r[col_a] != 0 else None, axis=1
    )
    merged = merged.sort_values('Δ выручка, тыс.', ascending=False).reset_index(drop=True)

    headers = list(merged.columns)
    ws.append(headers)
    _style_header(ws, 1)
    _write_df(ws, merged)

    delta_col_idx = headers.index('Δ выручка, тыс.') + 1
    pct_col_idx   = headers.index('Δ %') + 1
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row_idx, delta_col_idx).value
        if val is not None:
            font = POS_FONT if float(val) >= 0 else NEG_FONT
            ws.cell(row_idx, delta_col_idx).font = font
            ws.cell(row_idx, pct_col_idx).font = font

    _auto_width(ws)
    return merged


# ============================================================
# ЛИСТ 4: ДВИЖЕНИЕ КЛИЕНТОВ
# ============================================================

def _sheet_movement(wb, df_a, df_b, label_a, label_b):
    ws = wb.create_sheet("04_Движение_клиентов")

    clients_a = set(df_a['КОНЕЧНЫЙ_КЛИЕНТ'].dropna().unique())
    clients_b = set(df_b['КОНЕЧНЫЙ_КЛИЕНТ'].dropna().unique())

    new_clients  = clients_b - clients_a
    lost_clients = clients_a - clients_b
    kept_clients = clients_a & clients_b

    rev_b = df_b.groupby('КОНЕЧНЫЙ_КЛИЕНТ')[COL_REVENUE].sum()
    rev_a = df_a.groupby('КОНЕЧНЫЙ_КЛИЕНТ')[COL_REVENUE].sum()

    # Итоговая сводка
    summary_data = [
        ("Постоянные клиенты (были в обоих периодах)", len(kept_clients),
         round(rev_a[rev_a.index.isin(kept_clients)].sum() / 1000, 2),
         round(rev_b[rev_b.index.isin(kept_clients)].sum() / 1000, 2)),
        (f"Новые клиенты (только в {label_b})", len(new_clients),
         0,
         round(rev_b[rev_b.index.isin(new_clients)].sum() / 1000, 2)),
        (f"Ушедшие клиенты (только в {label_a})", len(lost_clients),
         round(rev_a[rev_a.index.isin(lost_clients)].sum() / 1000, 2),
         0),
    ]

    ws.append(["Категория", "Клиентов", f"Выручка {label_a}, тыс.", f"Выручка {label_b}, тыс."])
    _style_header(ws, 1)

    fills = [None, NEW_FILL, LOST_FILL]
    for i, row_data in enumerate(summary_data):
        ws.append(list(row_data))
        r = ws.max_row
        for c in range(1, 5):
            ws.cell(r, c).font = NEUTRAL_FONT
            ws.cell(r, c).border = THIN_BORDER
            if fills[i]:
                ws.cell(r, c).fill = fills[i]

    ws.append([])

    # Детальный список новых
    ws.append([f"НОВЫЕ КЛИЕНТЫ — появились в {label_b}"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, name="Segoe UI", size=10)
    ws.append(["Клиент", f"Выручка {label_b}, тыс.", "Заказов"])
    _style_header(ws, ws.max_row)

    new_detail = (
        df_b[df_b['КОНЕЧНЫЙ_КЛИЕНТ'].isin(new_clients)]
        .groupby('КОНЕЧНЫЙ_КЛИЕНТ')
        .agg(Выручка=(COL_REVENUE, 'sum'), Заказов=(COL_ORDER, 'count'))
        .reset_index()
        .sort_values('Выручка', ascending=False)
    )
    new_detail['Выручка'] = (new_detail['Выручка'] / 1000).round(2)
    for _, row in new_detail.iterrows():
        ws.append([row['КОНЕЧНЫЙ_КЛИЕНТ'], row['Выручка'], row['Заказов']])
        r = ws.max_row
        for c in range(1, 4):
            ws.cell(r, c).fill = NEW_FILL
            ws.cell(r, c).font = NEUTRAL_FONT
            ws.cell(r, c).border = THIN_BORDER

    ws.append([])

    # Детальный список ушедших
    ws.append([f"УШЕДШИЕ КЛИЕНТЫ — были в {label_a}, нет в {label_b}"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, name="Segoe UI", size=10)
    ws.append(["Клиент", f"Выручка {label_a}, тыс.", "Заказов"])
    _style_header(ws, ws.max_row)

    lost_detail = (
        df_a[df_a['КОНЕЧНЫЙ_КЛИЕНТ'].isin(lost_clients)]
        .groupby('КОНЕЧНЫЙ_КЛИЕНТ')
        .agg(Выручка=(COL_REVENUE, 'sum'), Заказов=(COL_ORDER, 'count'))
        .reset_index()
        .sort_values('Выручка', ascending=False)
    )
    lost_detail['Выручка'] = (lost_detail['Выручка'] / 1000).round(2)
    for _, row in lost_detail.iterrows():
        ws.append([row['КОНЕЧНЫЙ_КЛИЕНТ'], row['Выручка'], row['Заказов']])
        r = ws.max_row
        for c in range(1, 4):
            ws.cell(r, c).fill = LOST_FILL
            ws.cell(r, c).font = NEUTRAL_FONT
            ws.cell(r, c).border = THIN_BORDER

    _auto_width(ws)


# ============================================================
# ЛИСТ 5: ГРАФИКИ
# ============================================================

def _sheet_charts(wb, mgr_merged, cli_merged, label_a, label_b):
    ws = wb.create_sheet("05_Графики")
    ws['A1'] = "Сравнительные графики"
    ws['A1'].font = Font(bold=True, size=13, name="Segoe UI")

    col_a = f'Выручка_{label_a}_тыс'
    col_b = f'Выручка_{label_b}_тыс'

    # ── Данные для графика менеджеров ────────────────────────
    ws['A3'] = "Менеджер"
    ws['B3'] = label_a
    ws['C3'] = label_b
    for cell in [ws['A3'], ws['B3'], ws['C3']]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for i, row in mgr_merged.head(10).iterrows():
        r = 4 + i
        ws.cell(r, 1, row[COL_MANAGER])
        ws.cell(r, 2, float(row[col_a]) if col_a in mgr_merged.columns else 0)
        ws.cell(r, 3, float(row[col_b]) if col_b in mgr_merged.columns else 0)

    n_mgr = min(10, len(mgr_merged))
    chart_mgr = BarChart()
    chart_mgr.type = "bar"
    chart_mgr.title = f"Выручка по менеджерам: {label_a} vs {label_b}"
    chart_mgr.y_axis.title = "тыс. руб."
    chart_mgr.style = 10
    chart_mgr.width = 22
    chart_mgr.height = 14

    cats = Reference(ws, min_col=1, min_row=4, max_row=3 + n_mgr)
    data_mgr = Reference(ws, min_col=2, max_col=3, min_row=3, max_row=3 + n_mgr)
    chart_mgr.add_data(data_mgr, titles_from_data=True)
    chart_mgr.set_categories(cats)
    ws.add_chart(chart_mgr, "E3")

    # ── Данные для графика клиентов (топ-15) ─────────────────
    start_row = 4 + n_mgr + 3
    ws.cell(start_row, 1, "Клиент")
    ws.cell(start_row, 2, label_a)
    ws.cell(start_row, 3, label_b)
    for c in range(1, 4):
        ws.cell(start_row, c).fill = HEADER_FILL
        ws.cell(start_row, c).font = HEADER_FONT

    # Топ-15 по сумме двух периодов
    cli_top = cli_merged.copy()
    cli_top['_total'] = cli_top[col_a] + cli_top[col_b]
    cli_top = cli_top.nlargest(15, '_total')

    for i, (_, row) in enumerate(cli_top.iterrows()):
        r = start_row + 1 + i
        ws.cell(r, 1, row['КОНЕЧНЫЙ_КЛИЕНТ'])
        ws.cell(r, 2, float(row[col_a]) if col_a in cli_merged.columns else 0)
        ws.cell(r, 3, float(row[col_b]) if col_b in cli_merged.columns else 0)

    n_cli = len(cli_top)
    chart_cli = BarChart()
    chart_cli.type = "bar"
    chart_cli.title = f"Топ-15 клиентов: {label_a} vs {label_b}"
    chart_cli.y_axis.title = "тыс. руб."
    chart_cli.style = 10
    chart_cli.width = 22
    chart_cli.height = 16

    cats2 = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + n_cli)
    data_cli = Reference(ws, min_col=2, max_col=3, min_row=start_row, max_row=start_row + n_cli)
    chart_cli.add_data(data_cli, titles_from_data=True)
    chart_cli.set_categories(cats2)
    ws.add_chart(chart_cli, f"E{start_row}")

    _auto_width(ws)


# ============================================================
# ГЛАВНАЯ ФУНКЦИЯ
# ============================================================

def run_comparison(
    path_a: str,
    path_b: str,
    output_path: str,
    label_a: str = "Период А",
    label_b: str = "Период Б",
    log=print
):
    """
    Сравнивает два периода и сохраняет отчёт в output_path.
    path_a / path_b — пути к Excel-выгрузкам из CRM
    label_a / label_b — произвольные метки (например, "Янв 2024")
    """
    log(f"Загружаю период А: {os.path.basename(path_a)}")
    df_a = _load_period(path_a, log)

    log(f"Загружаю период Б: {os.path.basename(path_b)}")
    df_b = _load_period(path_b, log)

    log("Формирую отчёт...")

    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)  # удаляем дефолтный лист

    _sheet_summary(wb, df_a, df_b, label_a, label_b)
    mgr_merged = _sheet_managers(wb, df_a, df_b, label_a, label_b)
    cli_merged = _sheet_clients(wb, df_a, df_b, label_a, label_b)
    _sheet_movement(wb, df_a, df_b, label_a, label_b)

    if not mgr_merged.empty and not cli_merged.empty:
        _sheet_charts(wb, mgr_merged, cli_merged, label_a, label_b)

    wb.save(output_path)
    log(f"✅ Готово! Отчёт сравнения сохранён: {output_path}")
    return output_path
